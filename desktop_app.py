import sys
import os
import time
import subprocess
import requests
import ctypes
import math
from functools import partial
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QStackedWidget, QFrame,
    QMessageBox, QGridLayout, QComboBox, QDialog, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QFormLayout, QInputDialog, QCalendarWidget, QDialogButtonBox,
    QFileDialog, QToolButton, QMenu, QDateEdit
)
from PySide6.QtCore import Qt, QTimer, QDate
from PySide6.QtGui import QFont, QColor

API_URL = "http://127.0.0.1:8000"
MUTEX_NAME = "ShiftApp_Mutex_SingleInstance"


class ShiftApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.shifts_tab = None
        self.token = None
        self.user_role = None
        self.username = None
        self.server_process = None
        self.server_started = False
        self.blank_combo = None
        self.in_taken = None
        self.in_produced = None
        self.in_defect = None
        self.in_product = None
        self.in_reason = None
        self.prod_table = None

        self.setWindowTitle("Учёт смен и производства")
        self.setMinimumSize(900, 650)
        # Тема по умолчанию (светлая)
        self.current_theme = "light"

        # Загружаем сохранённую тему
        self.load_theme_preference()

        # Применяем тему
        self.apply_theme()

        # Пробуем запустить сервер в фоне
        QTimer.singleShot(100, self.try_start_server)

        self.init_ui()

    def get_light_theme(self):
        """Светлая тема (твоя оригинальная)"""
        return """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #667eea, stop:1 #764ba2);
            }
            QWidget { font-family: 'Segoe UI', Arial, sans-serif; }
            QLabel { color: #333; }

            QLineEdit, QComboBox {
                padding: 10px;
                border: 2px solid #89CFF0;
                border-radius: 8px;
                font-size: 14px;
                background-color: #E6F7FF;
                color: #004080;
            }

            QLineEdit:focus, QComboBox:focus {
                border: 2px solid #667eea;
                background-color: #FFFFFF;
                color: #333333;
            }

            QTabWidget::pane {
                border: 1px solid #ccc;
                background: white;
                border-radius: 10px;
            }

            QTabBar::tab {
                background-color: #E6F7FF;
                color: #004080;
                padding: 10px 20px;
                border: 1px solid #89CFF0;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 2px;
                font-weight: bold;
            }

            QTabBar::tab:selected {
                background: white;
                color: #667eea;
                border-bottom: 2px solid #667eea;
            }

            QTabBar::tab:hover:!selected {
                background: #D1EEFC;
            }

            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #667eea, stop:1 #764ba2);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #5568d3, stop:1 #65408b);
            }

            QFrame { background: white; border-radius: 15px; }
            QTableWidget { border: 1px solid #e0e0e0; border-radius: 5px; color: #333; background: white; }
            QHeaderView::section { background: #f8f9fa; padding: 5px; border: 1px solid #e0e0e0; color: #333; }
        """

    def get_dark_theme(self):
        """Тёмная тема"""
        return """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #16213e);
            }
            QWidget { font-family: 'Segoe UI', Arial, sans-serif; }
            QLabel { color: #e0e0e0; }

            QLineEdit, QComboBox {
                padding: 10px;
                border: 2px solid #4a5568;
                border-radius: 8px;
                font-size: 14px;
                background-color: #2d3748;
                color: #e0e0e0;
            }

            QLineEdit:focus, QComboBox:focus {
                border: 2px solid #667eea;
                background-color: #1a202c;
                color: #ffffff;
            }

            QTabWidget::pane {
                border: 1px solid #4a5568;
                background: #1a202c;
                border-radius: 10px;
            }

            QTabBar::tab {
                background-color: #2d3748;
                color: #a0aec0;
                padding: 10px 20px;
                border: 1px solid #4a5568;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 2px;
                font-weight: bold;
            }

            QTabBar::tab:selected {
                background: #1a202c;
                color: #667eea;
                border-bottom: 2px solid #667eea;
            }

            QTabBar::tab:hover:!selected {
                background: #4a5568;
                color: #e0e0e0;
            }

            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #667eea, stop:1 #764ba2);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #5568d3, stop:1 #65408b);
            }

            QFrame { background: #1a202c; border-radius: 15px; }

            /* Поля ввода внутри форм */
            QFrame QLineEdit,
            QFrame QComboBox {
                background: #2d3748;
                color: #e0e0e0;
                border: 2px solid #4a5568;
            }

            QFrame QLabel {
                color: #e0e0e0;
            }

            QTableWidget { border: 1px solid #4a5568; border-radius: 5px; color: #e0e0e0; background: #1a202c; }
            QHeaderView::section { background: #2d3748; padding: 5px; border: 1px solid #4a5568; color: #e0e0e0; }
        """

    def apply_theme(self):
        """Применяет текущую тему"""
        if self.current_theme == "dark":
            self.setStyleSheet(self.get_dark_theme())
        else:
            self.setStyleSheet(self.get_light_theme())

    def toggle_theme(self):
        """Переключает тему"""
        self.current_theme = "dark" if self.current_theme == "light" else "light"
        self.apply_theme()
        self.save_theme_preference()

        # Обновляем обе кнопки (на входе и в dashboard)
        icon = "" if self.current_theme == "light" else "☀️"

        if hasattr(self, 'theme_btn_login'):
            self.theme_btn_login.setText(icon)
            # Обновляем стиль
            if self.current_theme == "light":
                self.theme_btn_login.setStyleSheet("""
                    QPushButton {
                        background: #2d3748;
                        color: white;
                        border: 2px solid #4a5568;
                        border-radius: 25px;
                        font-size: 20px;
                    }
                    QPushButton:hover {
                        background: #1a202c;
                    }
                """)
            else:
                self.theme_btn_login.setStyleSheet("""
                    QPushButton {
                        background: #f6e05e;
                        color: #1a202c;
                        border: 2px solid #ecc94b;
                        border-radius: 25px;
                        font-size: 20px;
                    }
                    QPushButton:hover {
                        background: #ecc94b;
                    }
                """)

        if hasattr(self, 'theme_btn'):
            self.theme_btn.setText(icon)
            # Обновляем стиль
            if self.current_theme == "light":
                self.theme_btn.setStyleSheet("""
                    QPushButton {
                        background: #2d3748;
                        color: white;
                        border: 2px solid #4a5568;
                        border-radius: 25px;
                        font-size: 20px;
                    }
                    QPushButton:hover {
                        background: #1a202c;
                    }
                """)
            else:
                self.theme_btn.setStyleSheet("""
                    QPushButton {
                        background: #f6e05e;
                        color: #1a202c;
                        border: 2px solid #ecc94b;
                        border-radius: 25px;
                        font-size: 20px;
                    }
                    QPushButton:hover {
                        background: #ecc94b;
                    }
                """)

    def save_theme_preference(self):
        """Сохраняет выбор темы в файл"""
        try:
            with open("theme.txt", "w") as f:
                f.write(self.current_theme)
        except:
            pass

    def load_theme_preference(self):
        """Загружает сохранённую тему из файла"""
        try:
            with open("theme.txt", "r") as f:
                theme = f.read().strip()
                if theme in ["light", "dark"]:
                    self.current_theme = theme
        except:
            self.current_theme = "light"

    def _get_status_text(self, status):
        """Возвращает текст статуса с эмодзи"""
        mapping = {
            "working": "🟢 Работает",
            "late": "🔵 Опоздал",
            "absent": "🔴 Не вышел",
            "sick": "🟡 Болен",
            "vacation": "🟣 Отпуск",
            "no_shift": "⚪ Нет смены"
        }
        return mapping.get(status, status)

    def _get_status_color(self, status):
        from PySide6.QtGui import QColor
        colors = {
            "working": QColor("#28a745"),  # Насыщенный зелёный — работает вовремя
            "late": QColor("#fd7e14"),  # Оранжевый — опоздал
            "absent": QColor("#dc3545"),  # Красный — не вышел
            "sick": QColor("#ffc107"),  # Жёлтый — больничный
            "vacation": QColor("#ffc107"),  # Жёлтый — отпуск
            "no_shift": QColor("#6c757d")  # Серый — смена закрыта
        }
        return colors.get(status, QColor("white"))

    def _get_status_text(self, status):
        mapping = {
            "На работе": "На работе",
            "Опоздал": "Опоздал",
            "Не_вышел": "Не вышел",
            "Больничный": "Больничный",
            "В_отпуске": "В отпуске",
            "Смена_закрыта": "Смена закрыта"
        }
        return mapping.get(status, status)

    def try_start_server(self):
        """Пытаемся запустить сервер в фоне"""
        try:
            requests.get(f"{API_URL}/docs", timeout=1)
            self.server_started = True
            return
        except:
            pass

        if self.server_started:
            return

        self.server_started = True

        if getattr(sys, 'frozen', False):
            app_path = os.path.dirname(sys.executable)
            main_path = os.path.join(app_path, 'main.py')
            python_exe = r"C:\Users\Admin\AppData\Local\Programs\Python\Python313\python.exe"
        else:
            app_path = os.path.dirname(os.path.abspath(__file__))
            main_path = os.path.join(app_path, 'main.py')
            python_exe = r"C:\Users\Admin\AppData\Local\Programs\Python\Python313\python.exe"

        if not os.path.exists(main_path):
            return

        try:
            self.server_process = subprocess.Popen(
                [python_exe, "-m", "uvicorn", "main:app", "--host", "127.0.0.1",
                 "--port", "8000", "--log-level", "error"],
                cwd=app_path,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
            )
            QTimer.singleShot(3000, self.check_server_ready)
        except Exception as e:
            print(f"Ошибка запуска сервера: {e}")

    def check_server_ready(self):
        try:
            requests.get(f"{API_URL}/docs", timeout=1)
        except:
            pass

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.container = QFrame()
        self.container.setMinimumSize(800, 600)
        container_layout = QVBoxLayout(self.container)
        container_layout.setContentsMargins(20, 20, 20, 20)

        self.stack = QStackedWidget()
        self.stack.addWidget(self.create_login_page())
        self.stack.addWidget(self.create_register_page())
        self.stack.addWidget(self.create_dashboard_page())
        container_layout.addWidget(self.stack)
        main_layout.addWidget(self.container)

    def create_login_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(20)
        layout.addWidget(self._title_label("Учёт смен"))

        # Кнопка переключения темы (в правом верхнем углу)
        theme_container = QWidget()
        theme_layout = QHBoxLayout(theme_container)
        theme_layout.addStretch()

        self.theme_btn_login = QPushButton("" if self.current_theme == "light" else "☀️")
        self.theme_btn_login.setFixedSize(50, 50)

        # Адаптивный стиль
        if self.current_theme == "light":
            self.theme_btn_login.setStyleSheet("""
                QPushButton {
                    background: #2d3748;
                    color: white;
                    border: 2px solid #4a5568;
                    border-radius: 25px;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background: #1a202c;
                }
            """)
        else:
            self.theme_btn_login.setStyleSheet("""
                QPushButton {
                    background: #f6e05e;
                    color: #1a202c;
                    border: 2px solid #ecc94b;
                    border-radius: 25px;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background: #ecc94b;
                }
            """)

        self.theme_btn_login.clicked.connect(self.toggle_theme)
        theme_layout.addWidget(self.theme_btn_login)
        layout.addWidget(theme_container)

        layout.addWidget(QLabel("Логин"))
        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText("Введите логин")
        layout.addWidget(self.login_input)
        layout.addWidget(QLabel("Пароль"))
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)
        self.login_btn = QPushButton("Войти")
        self.login_btn.setMinimumHeight(50)
        self.login_btn.clicked.connect(self.login)
        layout.addWidget(self.login_btn)
        to_reg = QPushButton("Нет аккаунта? Зарегистрироваться")
        to_reg.setStyleSheet("background:transparent;color:#667eea;border:2px solid #667eea;")
        to_reg.clicked.connect(lambda: self.stack.setCurrentIndex(1))
        layout.addWidget(to_reg)
        layout.addStretch()
        return page

    def create_register_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(20)
        layout.addWidget(self._title_label("Регистрация"))
        layout.addWidget(QLabel("Логин"))
        self.reg_login = QLineEdit()
        self.reg_login.setPlaceholderText("Придумайте логин")
        layout.addWidget(self.reg_login)
        layout.addWidget(QLabel("Пароль"))
        self.reg_pass = QLineEdit()
        self.reg_pass.setPlaceholderText("Придумайте пароль")
        self.reg_pass.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.reg_pass)
        layout.addWidget(QLabel("Роль"))
        self.reg_role = QComboBox()
        self.reg_role.addItems(["Сотрудник", "Администратор", "Руководитель"])
        self.reg_role.currentTextChanged.connect(self._toggle_secret)
        layout.addWidget(self.reg_role)

        self.secret_widget = QWidget()
        sl = QVBoxLayout(self.secret_widget)
        sl.setContentsMargins(0, 0, 0, 0)
        sl.addWidget(QLabel("Секретный ключ админа"))
        self.reg_secret = QLineEdit()
        self.reg_secret.setPlaceholderText("SUPER_SECRET_ADMIN_KEY_2026")
        sl.addWidget(self.reg_secret)
        layout.addWidget(self.secret_widget)
        self.secret_widget.hide()

        reg_btn = QPushButton("Создать аккаунт")
        reg_btn.setMinimumHeight(50)
        reg_btn.setStyleSheet("background:qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #11998e,stop:1 #38ef7d);")
        reg_btn.clicked.connect(self.register)
        layout.addWidget(reg_btn)
        back_btn = QPushButton("Назад ко входу")
        back_btn.setStyleSheet("background:#6c757d;")
        back_btn.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        layout.addWidget(back_btn)
        layout.addStretch()
        return page

    def _on_tab_changed(self, index):
        """Вызывается при переключении вкладки"""
        if self.tabs.widget(index) == self.production_tab:
            print("🔄 Переключились на вкладку Производство")
            self.load_production_data()

    def create_dashboard_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(15)

        # Верхняя панель
        top_bar = QHBoxLayout()
        self.welcome = QLabel("Личный кабинет")
        self.welcome.setFont(QFont('Segoe UI', 18, QFont.Weight.Bold))
        top_bar.addWidget(self.welcome)
        top_bar.addStretch()

        # Кнопка переключения темы
        self.theme_btn = QPushButton("" if self.current_theme == "light" else "☀️")
        self.theme_btn.setFixedSize(50, 50)

        # Адаптивный стиль кнопки
        if self.current_theme == "light":
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background: #2d3748;
                    color: white;
                    border: 2px solid #4a5568;
                    border-radius: 25px;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background: #1a202c;
                }
            """)
        else:
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background: #f6e05e;
                    color: #1a202c;
                    border: 2px solid #ecc94b;
                    border-radius: 25px;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background: #ecc94b;
                }
            """)

        self.theme_btn.clicked.connect(self.toggle_theme)
        top_bar.addWidget(self.theme_btn)

        # Кнопка выхода
        out_btn = QPushButton("Выйти")
        out_btn.setFixedWidth(100)
        out_btn.setStyleSheet("background:#dc3545; color: white; border-radius: 8px; font-weight: bold;")
        out_btn.clicked.connect(self.logout)
        top_bar.addWidget(out_btn)

        layout.addLayout(top_bar)

        # Табы
        self.tabs = QTabWidget()

        # Вкладка 1: Смены (будет настроена при входе)
        self.shifts_tab = QWidget()
        self.tabs.addTab(self.shifts_tab, "Смены")

        # Вкладка 2: Производство
        self.production_tab = QWidget()
        self.setup_production_tab()
        self.tabs.addTab(self.production_tab, "Производство")

        # Загружаем данные при переключении на вкладку
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # Вкладка 3: Мониторинг (Только админ)
        self.monitoring_tab = QWidget()
        self.setup_monitoring_tab()

        # Вкладка 4: Склад сырья
        self.raw_tab = QWidget()
        self.setup_raw_tab()

        # Вкладка 5: Склад заготовок (Только админ)
        self.blanks_tab = QWidget()
        self.setup_blanks_tab()

        # Вкладка Финансы (Только Руководитель)
        self.finance_tab = QWidget()
        self.setup_finance_tab()

        # Вкладка Зарплата (Только Руководитель)
        self.payroll_tab = QWidget()
        self.setup_payroll_tab()

        # Вкладка 6: Склад заготовок (Только админ)
        self.blanks_tab = QWidget()
        self.setup_blanks_tab()
        self.tabs.addTab(self.blanks_tab, "Склад заготовок")

        # Вкладка Рецептуры (Только админ и руководитель)
        self.recipes_tab = QWidget()
        self.setup_recipes_tab()
        self.tabs.addTab(self.recipes_tab, "Рецептуры")

        # Вкладка График отпусков
        self.vacation_tab = QWidget()
        self.setup_vacation_tab()
        self.tabs.addTab(self.vacation_tab, "График отпусков")

        layout.addWidget(self.tabs)
        return page

    def load_vacation_schedule(self):
        """Загрузка графика отпусков всех сотрудников"""
        print(f"🔄 Загрузка графика отпусков...")

        try:
            # Получаем статус всех сотрудников на сегодня
            today = datetime.now().strftime("%Y-%m-%d")
            r = requests.get(f"{API_URL}/admin/monitoring/status?target_date={today}",
                             headers={"Authorization": f"Bearer {self.token}"})

            if r.status_code == 200:
                employees = r.json()
                self.vacation_table.setRowCount(len(employees))

                for i, emp in enumerate(employees):
                    print(f"  [{i}] {emp['login']}: {emp['status']}")

                    # Колонка 0: ID
                    self.vacation_table.setItem(i, 0, QTableWidgetItem(str(emp['id'])))
                    # Колонка 1: Сотрудник
                    self.vacation_table.setItem(i, 1, QTableWidgetItem(emp['login']))
                    # Колонка 2: Роль
                    self.vacation_table.setItem(i, 2, QTableWidgetItem(emp.get('role', 'user')))

                    # Колонка 3: Кнопка Изменить (только для админа и руководителя)
                    if self.user_role in ["admin", "manager"]:
                        edit_btn = QPushButton("✏️")
                        edit_btn.setToolTip("Добавить/Изменить отсутствие")
                        edit_btn.setFixedSize(30, 30)
                        edit_btn.setStyleSheet("""
                            QPushButton {
                                background: #28a745;
                                color: white;
                                border: none;
                                border-radius: 5px;
                                font-size: 14px;
                            }
                            QPushButton:hover {
                                background: #218838;
                            }
                        """)
                        edit_btn.clicked.connect(
                            lambda checked, uid=emp['id'], uname=emp['login']:
                            self.edit_absence_dialog(uid, uname)
                        )

                        self.vacation_table.setCellWidget(i, 3, edit_btn)
                    else:
                        self.vacation_table.setItem(i, 3, QTableWidgetItem("-"))

                # Скрываем колонку с ID
                self.vacation_table.setColumnHidden(0, True)

                print(f"✅ Загружено сотрудников: {len(employees)}")
            elif r.status_code == 403:
                print(f"⚠️ Нет доступа к мониторингу (403)")
                self.vacation_table.setRowCount(0)
            else:
                print(f"❌ Ошибка сервера: {r.status_code}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить данные: {r.status_code}")

        except Exception as e:
            print(f"❌ Исключение при загрузке: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def add_absence_dialog(self):
        """Диалог добавления отсутствия (отпуск/больничный)"""
        # Сначала выбираем сотрудника
        try:
            r = requests.get(f"{API_URL}/manager/users/list",
                             headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code != 200:
                QMessageBox.critical(self, "Ошибка", "Не удалось загрузить список сотрудников")
                return

            users = r.json()
            user_names = [f"{u['login']} ({u['role']})" for u in users]
            user_ids = [u['id'] for u in users]

            user_name, ok = QInputDialog.getItem(
                self, "Выбор сотрудника", "Выберите сотрудника:", user_names, 0, False
            )

            if not ok or not user_name:
                return

            # Получаем ID выбранного пользователя
            selected_index = user_names.index(user_name)
            user_id = user_ids[selected_index]

            # Теперь показываем диалог с датами
            self._show_absence_date_dialog(user_id, user_name.split()[0])

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def edit_absence_dialog(self, user_id, user_name):
        """Диалог редактирования отсутствия для конкретного сотрудника"""
        self._show_absence_date_dialog(user_id, user_name)

    def _show_absence_date_dialog(self, user_id, user_name):
        """Диалог выбора дат и типа отсутствия"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Отсутствие: {user_name}")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Тип отсутствия
        type_combo = QComboBox()
        type_combo.addItems(["Больничный", "Отпуск"])
        layout.addRow("Тип:", type_combo)

        # Дата начала
        start_date = QDateEdit()
        start_date.setDate(QDate.currentDate())
        start_date.setCalendarPopup(True)
        start_date.setDisplayFormat("yyyy-MM-dd")
        layout.addRow("Дата начала:", start_date)

        # Дата конца
        end_date = QDateEdit()
        end_date.setDate(QDate.currentDate().addDays(14))
        end_date.setCalendarPopup(True)
        end_date.setDisplayFormat("yyyy-MM-dd")
        layout.addRow("Дата конца:", end_date)

        # Комментарий
        comment_input = QLineEdit()
        comment_input.setPlaceholderText("Комментарий (необязательно)")
        layout.addRow("Комментарий:", comment_input)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            absence_type = "sick" if type_combo.currentText() == "Больничный" else "vacation"
            start = start_date.date().toString("yyyy-MM-dd")
            end = end_date.date().toString("yyyy-MM-dd")
            comment = comment_input.text().strip()

            try:
                payload = {
                    "user_id": int(user_id),
                    "start_date": start,
                    "end_date": end,
                    "absence_type": absence_type,
                    "comment": comment
                }

                r = requests.post(f"{API_URL}/admin/monitoring/set_absence",
                                  json=payload,
                                  headers={"Authorization": f"Bearer {self.token}"})

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех",
                                            f"Отсутствие добавлено!\n{user_name}: {absence_type} с {start} по {end}")
                    self.load_vacation_schedule()
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка добавления"))

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def delete_absence_dialog(self, user_id, user_name):
        """Диалог удаления отсутствия"""
        reply = QMessageBox.question(
            self, "Удаление отсутствия",
            f"Вы уверены, что хотите удалить все отсутствия для '{user_name}'?\nЭто действие необратимо!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Удаляем все отсутствия пользователя
                payload = {
                    "user_id": int(user_id),
                    "start_date": "2020-01-01",
                    "end_date": "2030-12-31",
                    "absence_type": None
                }

                r = requests.post(f"{API_URL}/admin/monitoring/set_absence",
                                  json=payload,
                                  headers={"Authorization": f"Bearer {self.token}"})

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Все отсутствия для '{user_name}' удалены!")
                    self.load_vacation_schedule()
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка удаления"))

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def setup_vacation_tab(self):
        layout = QVBoxLayout(self.vacation_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)

        # Заголовок
        layout.addWidget(QLabel("<b>График отпусков и больничных</b>"))

        # Панель управления (только для админа и руководителя)
        if self.user_role in ["admin", "manager"]:
            controls = QHBoxLayout()

            add_btn = QPushButton("➕ Добавить отсутствие")
            add_btn.setStyleSheet("background: #28a745; color: white; border-radius: 8px; font-weight: bold;")
            add_btn.setMinimumHeight(40)
            add_btn.clicked.connect(self.add_absence_dialog)
            controls.addWidget(add_btn)

            controls.addStretch()
            layout.addLayout(controls)

        # Таблица сотрудников - 4 колонки (ID скрыт + Сотрудник + Роль + Действия)
        self.vacation_table = QTableWidget()
        self.vacation_table.setColumnCount(4)
        self.vacation_table.setHorizontalHeaderLabels(["ID", "Сотрудник", "Роль", "Действия"])
        self.vacation_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.vacation_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.vacation_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        layout.addWidget(self.vacation_table)

        # Кнопка обновления
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px; font-weight: bold;")
        refresh_btn.setMinimumHeight(40)
        refresh_btn.clicked.connect(self.load_vacation_schedule)
        layout.addWidget(refresh_btn)

    def reset_password_dialog(self, user_id, login):
        """Диалог смены пароля сотруднику"""
        new_pass, ok = QInputDialog.getText(
            self, "Смена пароля", f"Введите новый пароль для '{login}':",
            QLineEdit.EchoMode.Password
        )

        if ok and new_pass:
            try:
                payload = {"new_password": new_pass}
                r = requests.put(
                    f"{API_URL}/manager/users/{user_id}/password",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )
                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Пароль для {login} изменен!")
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def delete_user_dialog(self, user_id, login):
        """Диалог удаления сотрудника"""
        reply = QMessageBox.question(
            self, "Удаление",
            f"Вы уверены, что хотите удалить сотрудника '{login}'?\nЭто действие необратимо.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                r = requests.delete(
                    f"{API_URL}/manager/users/{user_id}",
                    headers={"Authorization": f"Bearer {self.token}"}
                )
                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Сотрудник {login} удален.")
                    self.load_finance_data()  # Обновляем список
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def start_shift(self):
        try:
            # ИСПРАВЛЕНО: вызываем /shifts/open вместо /shifts/start
            r = requests.post(f"{API_URL}/shifts/open", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех", "Смена начата успешно!")
                self.load_production_data()  # Обновить данные при старте
            else:
                QMessageBox.warning(self, "Ошибка", r.json().get("detail", "Не удалось начать смену"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def end_shift(self):
        try:
            # ИСПРАВЛЕНО: вызываем /shifts/close вместо /shifts/end
            r = requests.post(f"{API_URL}/shifts/close", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                dur = r.json().get("duration_hours", 0)
                QMessageBox.information(self, "Успех", f"Смена завершена! Отработано: {dur} ч.")
                self.load_production_data()  # Обновить данные
            else:
                QMessageBox.warning(self, "Ошибка", r.json().get("detail", "Не удалось завершить смену"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def show_shift_history(self):
        """Показывает историю смен текущего пользователя"""
        try:
            if self.user_role == "admin":
                url = f"{API_URL}/admin/shifts/all-history?limit=100"
            else:
                url = f"{API_URL}/shifts/history?limit=100"
            r = requests.get(url, headers={"Authorization": f"Bearer {self.token}"})

            if r.status_code != 200:
                QMessageBox.warning(self, "Ошибка", r.json().get("detail", "Не удалось загрузить историю"))
                return

            shifts = r.json()

            if not shifts:
                QMessageBox.information(self, "История", "У вас пока нет смен.")
                return

            # Формируем красивое сообщение
            msg = "<b>История ваших смен:</b><br><br>"

            for i, s in enumerate(shifts[:20], 1):  # Показываем последние 20
                # Парсим время начала
                start_str = s.get("start_time", "")
                if start_str:
                    try:
                        st = datetime.fromisoformat(start_str.replace("Z", "+00:00"))
                        start_formatted = st.strftime("%d.%m.%Y %H:%M")
                    except:
                        start_formatted = start_str
                else:
                    start_formatted = "неизвестно"

                # Парсим время конца
                end_str = s.get("end_time")
                if end_str:
                    try:
                        et = datetime.fromisoformat(end_str.replace("Z", "+00:00"))
                        end_formatted = et.strftime("%H:%M")
                    except:
                        end_formatted = end_str
                    duration = s.get("duration_hours", 0) or 0
                    time_info = f" до {end_formatted} ({duration:.1f} ч)"
                else:
                    time_info = " — <b>открыта</b>"

                # Опоздание
                late_info = ""
                if s.get("is_late"):
                    late_minutes = s.get("late_minutes", 0)
                    late_info = f" ⚠️ Опоздание: +{late_minutes} мин"

                msg += f"<b>{i}.</b> {start_formatted}{time_info}{late_info}<br>"

            # Показываем в красивом окне
            dialog = QDialog(self)
            dialog.setWindowTitle("📋 История смен")
            dialog.setMinimumSize(500, 400)

            layout = QVBoxLayout(dialog)

            label = QLabel(msg)
            label.setWordWrap(True)
            label.setStyleSheet("font-size: 13px; padding: 10px;")
            layout.addWidget(label)

            close_btn = QPushButton("Закрыть")
            close_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px; padding: 8px;")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)

            dialog.exec()

        except requests.exceptions.ConnectionError:
            QMessageBox.critical(self, "Ошибка", "Нет связи с сервером.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить историю: {str(e)}")

    def _setup_shifts_interface(self):
        """Создает интерфейс вкладки 'Смены' с нуля"""

        if not hasattr(self, 'shifts_tab') or not self.shifts_tab:
            print("❌ self.shifts_tab не существует!")
            return

        print(f"🛠️ Создаем интерфейс смен. Роль: {self.user_role}")

        # Проверяем, есть ли уже layout - если да, удаляем его полностью
        if self.shifts_tab.layout():
            # Принудительно удаляем все виджеты
            while self.shifts_tab.layout().count():
                item = self.shifts_tab.layout().takeAt(0)
                if item.widget():
                    item.widget().close()
                    item.widget().setParent(None)
            # Удаляем сам layout
            old_layout = self.shifts_tab.layout()
            self.shifts_tab.setLayout(None)
            del old_layout

        # Создаем НОВЫЙ layout
        new_layout = QVBoxLayout()
        new_layout.setContentsMargins(20, 20, 20, 20)
        new_layout.setSpacing(15)

        # === БЛОК 1: КНОПКИ СМЕНЫ (Для Сотрудника И Админа) ===
        if self.user_role in ["user", "admin"]:
            print("✅ Добавляем кнопки смен")

            # Статистика
            stats_layout = QHBoxLayout()
            self.total_shifts_label = QLabel("0")
            self.total_shifts_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #667eea;")
            self.total_shifts_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

            self.total_hours_label = QLabel("0.0")
            self.total_hours_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #667eea;")
            self.total_hours_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

            stats_layout.addWidget(QLabel("Всего смен"))
            stats_layout.addWidget(self.total_shifts_label)
            stats_layout.addWidget(QLabel("Часов"))
            stats_layout.addWidget(self.total_hours_label)
            new_layout.addLayout(stats_layout)

            # Кнопка "Начать" - ЗЕЛЁНАЯ
            self.start_shift_btn = QPushButton("Начать смену")
            self.start_shift_btn.setStyleSheet(
                "background: #28a745; color: white; border: none; border-radius: 8px; padding: 12px; font-size: 14px; font-weight: bold;"
            )
            self.start_shift_btn.clicked.connect(self.start_shift)
            new_layout.addWidget(self.start_shift_btn)

            # Кнопка "Завершить" - КРАСНАЯ
            self.end_shift_btn = QPushButton("Завершить смену")
            self.end_shift_btn.setStyleSheet(
                "background: #dc3545; color: white; border: none; border-radius: 8px; padding: 12px; font-size: 14px; font-weight: bold;"
            )
            self.end_shift_btn.clicked.connect(self.end_shift)
            new_layout.addWidget(self.end_shift_btn)

            # Кнопка "История" - СИНЯЯ
            self.history_btn = QPushButton("История смен")
            self.history_btn.setStyleSheet(
                "background: #007bff; color: white; border: none; border-radius: 8px; padding: 12px; font-size: 14px;"
            )
            self.history_btn.clicked.connect(self.show_shift_history)
            new_layout.addWidget(self.history_btn)

            if self.user_role == "admin":
                separator = QLabel("<hr>")
                separator.setStyleSheet("color: #ccc;")
                new_layout.addWidget(separator)

        # === БЛОК 2: МОНИТОРИНГ (Для Руководителя И Админа) ===
        if self.user_role in ["manager", "admin"]:
            print("✅ Добавляем кнопку мониторинга")

            monitoring_btn = QPushButton("📊 Мониторинг сотрудников")
            monitoring_btn.setStyleSheet("""
                QPushButton {
                    background: #6f42c1;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 15px;
                    font-size: 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background: #5a32a3;
                }
            """)
            monitoring_btn.clicked.connect(self.show_all_shifts_monitoring)
            new_layout.addWidget(monitoring_btn)

        new_layout.addStretch()

        # Устанавливаем layout
        self.shifts_tab.setLayout(new_layout)

        # Принудительно обновляем
        self.shifts_tab.update()
        self.shifts_tab.repaint()
        QApplication.processEvents()

        print("✅ Интерфейс смен создан")

    def setup_production_tab(self):
        layout = QVBoxLayout(self.production_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)

        # Форма отчёта
        form_box = QFrame()
        form_box.setObjectName("production_form")
        form_box.setStyleSheet("""
            QFrame#production_form {
                background: #f8f9fa;
                border-radius: 10px;
                padding: 20px;
            }
            QFrame#production_form QLabel {
                color: #004080;
                font-weight: bold;
            }
            QFrame#production_form QLineEdit,
            QFrame#production_form QComboBox {
                background: white;
                color: #333;
                border: 2px solid #89CFF0;
                border-radius: 8px;
                padding: 10px;
            }
        """)

        # Создаём grid layout для формы
        form_layout = QGridLayout(form_box)
        assert isinstance(form_layout, QGridLayout)  # Помощь IDE
        form_layout.setSpacing(12)
        form_layout.setContentsMargins(10, 10, 10, 10)
        form_layout.setColumnStretch(1, 1)

        # === СОЗДАЁМ ВИДЖЕТЫ ===
        self.blank_combo = QComboBox()
        self.blank_combo.setMinimumHeight(36)
        self.blank_combo.addItem("Загрузка заготовок...")

        self.in_taken = QLineEdit("1")
        self.in_taken.setMinimumHeight(36)
        self.in_produced = QLineEdit("0")
        self.in_produced.setMinimumHeight(36)
        self.in_defect = QLineEdit("0")
        self.in_defect.setMinimumHeight(36)

        self.in_reason = QLineEdit()
        self.in_reason.setMinimumHeight(36)
        self.in_reason.setPlaceholderText("Причина недостачи или брака")

        # Инициализируем список ID заготовок
        self.blank_ids = []

        # Стиль для подписей
        lbl_style = "font-weight: bold; color: #004080; font-size: 14px;"
        l_blank = QLabel("Заготовка:")
        l_blank.setStyleSheet(lbl_style)
        l_taken = QLabel("Взято заготовок (шт):")
        l_taken.setStyleSheet(lbl_style)
        l_prod = QLabel("Сделано годных (шт):")
        l_prod.setStyleSheet(lbl_style)
        l_defect = QLabel("Брак (шт):")
        l_defect.setStyleSheet(lbl_style)
        l_reason = QLabel("Причина:")
        l_reason.setStyleSheet(lbl_style)

        # === РАССТАВЛЯЕМ В СЕТКУ ===
        form_layout.addWidget(l_blank, 0, 0, 1, 1)
        form_layout.addWidget(self.blank_combo, 0, 1, 1, 1)

        form_layout.addWidget(l_taken, 1, 0, 1, 1)
        form_layout.addWidget(self.in_taken, 1, 1, 1, 1)

        form_layout.addWidget(l_prod, 2, 0, 1, 1)
        form_layout.addWidget(self.in_produced, 2, 1, 1, 1)

        form_layout.addWidget(l_defect, 3, 0, 1, 1)
        form_layout.addWidget(self.in_defect, 3, 1, 1, 1)

        form_layout.addWidget(l_reason, 4, 0, 1, 1)
        form_layout.addWidget(self.in_reason, 4, 1, 1, 1)

        # Кнопка на всю ширину формы
        submit_btn = QPushButton("Сдать отчёт")
        submit_btn.setStyleSheet(
            "background: #28a745; color: white; border-radius: 8px; font-weight: bold; font-size: 16px; padding: 8px;")
        submit_btn.setMinimumHeight(42)
        submit_btn.clicked.connect(self.submit_production_report)
        form_layout.addWidget(submit_btn, 5, 0, 1, 2)

        layout.addWidget(form_box)

        # Таблица истории
        layout.addWidget(QLabel("<b>История выработки:</b>"))
        self.prod_table = QTableWidget()
        self.prod_table.setColumnCount(5)
        self.prod_table.setHorizontalHeaderLabels(["Дата", "Изделие", "Взято", "Годных", "Брак"])
        self.prod_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.prod_table.verticalHeader().setVisible(False)
        layout.addWidget(self.prod_table)

        refresh_btn = QPushButton("Обновить данные")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px; font-weight: bold;")
        refresh_btn.setMinimumHeight(40)
        refresh_btn.clicked.connect(self.load_production_data)
        layout.addWidget(refresh_btn)

    def setup_blanks_tab(self):
        layout = QVBoxLayout(self.blanks_tab)
        layout.setSpacing(15)

        # Таблица - 5 колонок (ID скрыт + Название + Остаток + Действия)
        self.blanks_table = QTableWidget()
        self.blanks_table.setColumnCount(5)
        self.blanks_table.setHorizontalHeaderLabels(["ID", "Название", "Остаток", "", "Действия"])
        self.blanks_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.blanks_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.blanks_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        layout.addWidget(self.blanks_table)

        # Панель управления (убрана кнопка "Взять заготовки")
        controls = QHBoxLayout()

        self.blank_add_btn = QPushButton("➕ Добавить заготовки")
        self.blank_add_btn.setStyleSheet("background: #28a745; color: white; border-radius: 8px; font-weight: bold;")
        self.blank_add_btn.setMinimumHeight(40)
        self.blank_add_btn.clicked.connect(self.add_blank_dialog)

        self.blank_refresh_btn = QPushButton("🔄 Обновить")
        self.blank_refresh_btn.setStyleSheet(
            "background: #6c757d; color: white; border-radius: 8px; font-weight: bold;")
        self.blank_refresh_btn.setMinimumHeight(40)
        self.blank_refresh_btn.clicked.connect(self.load_blanks_list)

        controls.addWidget(self.blank_add_btn)
        controls.addStretch()  # Растягиваем пространство
        controls.addWidget(self.blank_refresh_btn)
        layout.addLayout(controls)

    def setup_raw_tab(self):
        layout = QVBoxLayout(self.raw_tab)
        layout.setSpacing(15)

        # Заголовок
        layout.addWidget(QLabel("<b>Склад сырья:</b>"))

        # Кнопка добавления сырья - ДЛЯ РУКОВОДИТЕЛЯ И АДМИНА
        if self.user_role in ["manager", "admin"]:
            add_raw_btn = QPushButton("➕ Добавить сырье (Приход)")
            add_raw_btn.setStyleSheet(
                "background: #20c997; color: white; border: none; "
                "border-radius: 8px; padding: 10px; font-weight: bold;"
            )
            add_raw_btn.clicked.connect(self.add_raw_material_dialog)
            layout.addWidget(add_raw_btn)

        # Таблица сырья - 6 колонок (ID скрыт + 4 видимых + Действия)
        self.raw_table = QTableWidget()
        self.raw_table.setColumnCount(6)
        self.raw_table.setHorizontalHeaderLabels(["ID", "Название", "Толщина", "Цвет", "Остаток", "Действия"])
        self.raw_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.raw_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.raw_table)

        # Кнопка обновления
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px;")
        refresh_btn.clicked.connect(self.load_raw_materials)
        layout.addWidget(refresh_btn)

    def load_raw_materials(self):
        print(f"🔄 Загрузка сырья для роли: {self.user_role}")
        print(f" Token: {'Есть' if self.token else 'НЕТ!'}")

        try:
            r = requests.get(f"{API_URL}/admin/warehouse/raw", headers={"Authorization": f"Bearer {self.token}"})
            print(f"📡 Статус ответа: {r.status_code}")

            if r.status_code == 200:
                materials = r.json()
                print(f"✅ Получено материалов: {len(materials)}")

                self.raw_table.setRowCount(len(materials))
                for i, mat in enumerate(materials):
                    print(
                        f"  [{i}] ID={mat['id']}, Name={mat['name']}, Thickness={mat['thickness']}, Color={mat['color']}, Qty={mat['quantity']}")
                    self.raw_table.setItem(i, 0, QTableWidgetItem(str(mat['id'])))
                    self.raw_table.setItem(i, 1, QTableWidgetItem(mat['name']))
                    self.raw_table.setItem(i, 2, QTableWidgetItem(mat['thickness']))
                    self.raw_table.setItem(i, 3, QTableWidgetItem(mat['color']))
                    self.raw_table.setItem(i, 4, QTableWidgetItem(str(mat['quantity'])))

                    # Кнопка с меню действий (колонка 5)
                    if self.user_role in ["admin", "manager"]:
                        action_btn = QPushButton("Действия ▼")
                        action_btn.setStyleSheet("""
                            QPushButton {
                                background: #6f42c1;
                                color: white;
                                border: none;
                                border-radius: 5px;
                                padding: 5px 10px;
                                font-size: 12px;
                                font-weight: bold;
                            }
                            QPushButton:hover {
                                background: #5a32a3;
                            }
                        """)

                        # Создаём меню
                        menu = QMenu()
                        menu.addAction("✏️ Изменить")
                        menu.addAction("🗑️ Удалить")

                        # Для отладки
                        print(f"  [{i}] Создаю кнопку для материала: ID={mat['id']}, Name={mat['name']}")

                        # Подключаем кнопку к показу меню
                        def show_action_menu(checked, mid=mat['id'], mname=mat['name'],
                                             mthickness=mat['thickness'], mcolor=mat['color'],
                                             mqty=mat['quantity']):
                            print(f"📍 Показываю меню для: ID={mid}, Name={mname}")
                            action = menu.exec(action_btn.mapToGlobal(action_btn.rect().bottomLeft()))
                            if action:
                                if action.text() == "✏️ Изменить":
                                    print(f"✏️ Выбрано 'Изменить'")
                                    self.edit_raw_material(mid, mname, mthickness, mcolor, mqty)
                                elif action.text() == "🗑️ Удалить":
                                    print(f"🗑️ Выбрано 'Удалить'")
                                    self.delete_raw_material(mid, mname)

                        action_btn.clicked.connect(show_action_menu)
                        action_btn.setToolTip("Нажмите для выбора действия")

                        self.raw_table.setCellWidget(i, 5, action_btn)
                        print(f"  ✅ Кнопка добавлена")
                    else:
                        self.raw_table.setItem(i, 5, QTableWidgetItem("-"))

                # Скрываем колонку с ID
                self.raw_table.setColumnHidden(0, True)
                print("✅ Колонка ID скрыта")
            else:
                print(f"❌ Ошибка сервера: {r.status_code}")
                print(f"   Ответ: {r.text}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить сырье: {r.status_code}")

        except Exception as e:
            print(f"❌ Исключение при загрузке: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def delete_raw_material(self, material_id, material_name):
        """Удаление материала со склада"""
        print(f"🗑️ Удаление материала ID={material_id}, Name={material_name}")
        reply = QMessageBox.question(
            self,
            "Удаление материала",
            f"Вы уверены, что хотите удалить '{material_name}'?\nЭто действие необратимо!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                r = requests.delete(
                    f"{API_URL}/admin/warehouse/raw/{material_id}",
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Материал '{material_name}' удален!")
                    self.load_raw_materials()  # Обновляем таблицу
                else:
                    error_msg = r.json().get("detail", "Ошибка удаления")
                    print(f"❌ Ошибка удаления: {error_msg}")
                    QMessageBox.critical(self, "Ошибка", error_msg)

            except Exception as e:
                print(f"❌ Исключение при удалении: {e}")
                QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def edit_raw_material(self, material_id, name, thickness, color, quantity):
        """Редактирование материала"""
        print(f"✏️ Редактирование: ID={material_id}, Name={name}, Thickness={thickness}, Color={color}, Qty={quantity}")

        dialog = QDialog(self)
        dialog.setWindowTitle("Изменить материал")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Название
        name_input = QLineEdit(name)
        name_input.setPlaceholderText("Название")
        layout.addRow("Название:", name_input)

        # Толщина
        thickness_input = QLineEdit(thickness)
        thickness_input.setPlaceholderText("Толщина (например: 3 мм)")
        layout.addRow("Толщина:", thickness_input)

        # Цвет
        color_input = QLineEdit(color)
        color_input.setPlaceholderText("Цвет")
        layout.addRow("Цвет:", color_input)

        # Количество
        qty_input = QLineEdit(str(quantity))
        qty_input.setPlaceholderText("Количество")
        layout.addRow("Количество:", qty_input)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_name = name_input.text().strip().lower()
            new_thickness = thickness_input.text().strip()
            new_color = color_input.text().strip()
            qty_text = qty_input.text().strip()

            if not new_name or not new_thickness or not new_color or not qty_text:
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                new_qty = float(qty_text)
                if new_qty < 0:
                    raise ValueError

                payload = {
                    "name": new_name,
                    "thickness": new_thickness,
                    "color": new_color,
                    "quantity": new_qty
                }

                print(f"📤 Отправка обновления: {payload}")
                r = requests.put(
                    f"{API_URL}/admin/warehouse/raw/{material_id}",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Материал '{new_name}' обновлен!")
                    self.load_raw_materials()
                else:
                    error_msg = r.json().get("detail", "Ошибка обновления")
                    print(f"❌ Ошибка обновления: {error_msg}")
                    QMessageBox.critical(self, "Ошибка", error_msg)

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Неверное количество!")
            except Exception as e:
                print(f"❌ Исключение при обновлении: {e}")
                QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def setup_monitoring_tab(self):
        layout = QVBoxLayout(self.monitoring_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)

        # Панель выбора даты
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Дата мониторинга:"))
        self.monitor_date_edit = QLineEdit(datetime.now().strftime("%Y-%m-%d"))
        self.monitor_date_edit.setMaximumWidth(150)
        date_layout.addWidget(self.monitor_date_edit)

        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px; font-weight: bold;")
        refresh_btn.clicked.connect(self.load_monitoring_status)
        date_layout.addWidget(refresh_btn)
        date_layout.addStretch()
        layout.addLayout(date_layout)

        # Таблица сотрудников
        self.monitor_table = QTableWidget()
        self.monitor_table.setColumnCount(3)
        self.monitor_table.setHorizontalHeaderLabels(["ID", "Сотрудник", "Статус"])
        self.monitor_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.monitor_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.monitor_table.itemDoubleClicked.connect(self.open_calendar_dialog)  # Двойной клик открывает календарь
        layout.addWidget(self.monitor_table)

        # Легенда
        legend = QLabel("<b>Легенда:</b> 🟢 На работе | 🔵 Опоздал | 🔴 Не вышел | 🟡 Болен | 🟣 Отпуск")
        legend.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(legend)


    def load_monitoring_status(self):
        target_date = self.monitor_date_edit.text()
        try:
            r = requests.get(f"{API_URL}/admin/monitoring/status?target_date={target_date}",
                             headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                data = r.json()
                self.monitor_table.setRowCount(len(data))

                for i, emp in enumerate(data):
                    self.monitor_table.setItem(i, 0, QTableWidgetItem(str(emp['id'])))
                    self.monitor_table.setItem(i, 1, QTableWidgetItem(emp['login']))

                    status_item = QTableWidgetItem(self._get_status_text(emp['status']))
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    status_item.setBackground(self._get_status_color(emp['status']))

                    status_item.setForeground(QColor("#000000"))

                    self.monitor_table.setItem(i, 2, status_item)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить данные: {e}")

    def _get_status_text(self, status):
        mapping = {
            "working": "На работе",
            "late": "Опоздал",
            "absent": "Не вышел",
            "sick": "Болен",
            "vacation": "Отпуск"
        }
        return mapping.get(status, status)

    def _get_status_color(self, status):
        from PySide6.QtGui import QColor
        colors = {
            "working": QColor("#28a745"),  # Насыщенный зелёный
            "late": QColor("#007bff"),  # Насыщенный синий
            "absent": QColor("#dc3545"),  # Насыщенный красный
            "sick": QColor("#ffc107"),  # Насыщенный жёлтый
            "vacation": QColor("#6f42c1"),  # Насыщенный фиолетовый
            "no_shift": QColor("#6c757d")  # Серый
        }
        return colors.get(status, QColor("white"))

    def open_calendar_dialog(self, item):
        row = self.monitor_table.currentRow()
        if row < 0: return

        user_id = self.monitor_table.item(row, 0).text()
        user_login = self.monitor_table.item(row, 1).text()

        dialog = QDialog(self)
        dialog.setWindowTitle(f"График: {user_login}")
        dialog.setMinimumSize(450, 450)

        layout = QVBoxLayout(dialog)

        # Календарь с множественным выбором
        calendar = QCalendarWidget()
        calendar.setGridVisible(True)
        calendar.setSelectionMode(QCalendarWidget.SelectionMode.SingleSelection)
        layout.addWidget(calendar)

        # Информация о выбранном периоде
        self.period_label = QLabel("Выберите начало периода")
        self.period_label.setStyleSheet("font-weight: bold; color: #004080;")
        layout.addWidget(self.period_label)

        # Кнопки
        btn_layout = QHBoxLayout()

        self.btn_start = QPushButton("📅 Начало периода")
        self.btn_start.clicked.connect(lambda: self._set_calendar_period_start(calendar))
        btn_layout.addWidget(self.btn_start)

        self.btn_end = QPushButton("📅 Конец периода")
        self.btn_end.clicked.connect(lambda: self._set_calendar_period_end(calendar))
        btn_layout.addWidget(self.btn_end)

        layout.addLayout(btn_layout)

        # Кнопки сохранения
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)

        # Сохраняем данные для доступа из других методов
        self.calendar_dialog_data = {
            'user_id': user_id,
            'user_login': user_login,
            'calendar': calendar,
            'period_start': None,
            'period_end': None
        }

        # Загружаем существующие отсутствия
        self._load_user_absences_to_calendar(calendar, user_id)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            if self.calendar_dialog_data['period_start'] and self.calendar_dialog_data['period_end']:
                self._show_absence_options(
                    user_id,
                    self.calendar_dialog_data['period_start'],
                    self.calendar_dialog_data['period_end']
                )
            else:
                QMessageBox.warning(self, "Внимание", "Выберите период (начало и конец)!")

    def _set_calendar_period_start(self, calendar):
        self.calendar_dialog_data['period_start'] = calendar.selectedDate().toPython()
        start_str = self.calendar_dialog_data['period_start'].strftime("%d.%m.%Y")
        end_str = self.calendar_dialog_data['period_end'].strftime("%d.%m.%Y") if self.calendar_dialog_data[
            'period_end'] else "..."
        self.period_label.setText(f"Период: {start_str} - {end_str}")

    def _set_calendar_period_end(self, calendar):
        self.calendar_dialog_data['period_end'] = calendar.selectedDate().toPython()
        start_str = self.calendar_dialog_data['period_start'].strftime("%d.%m.%Y") if self.calendar_dialog_data[
            'period_start'] else "..."
        end_str = self.calendar_dialog_data['period_end'].strftime("%d.%m.%Y")
        self.period_label.setText(f"Период: {start_str} - {end_str}")

    def _load_user_absences_to_calendar(self, calendar, user_id):
        try:
            year = datetime.now().year
            r = requests.get(f"{API_URL}/admin/monitoring/absences/{user_id}?year={year}",
                             headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                absences = r.json()
                # Здесь можно добавить визуальное выделение дней на календаре
                # Но QCalendarWidget не поддерживает нативное выделение диапазонов
                # Поэтому просто покажем в label
                if absences:
                    info_text = "\n".join([
                        f"{a['absence_type']}: {a['start_date']} - {a['end_date']}"
                        for a in absences
                    ])
                    # Можно добавить дополнительную метку с информацией
        except:
            pass

    def _show_absence_options(self, user_id, start_date, end_date):
        options = ["Убрать статус", "Больничный", "Отпуск"]
        choice, ok = QInputDialog.getItem(
            self,
            "Изменить статус",
            f"Выберите статус на период {start_date.strftime('%d.%m')} - {end_date.strftime('%d.%m.%Y')}:",
            options, 0, False
        )

        if not ok: return

        absence_map = {
            "Убрать статус": None,
            "Больничный": "sick",
            "Отпуск": "vacation"
        }

        payload = {
            "user_id": int(user_id),
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "absence_type": absence_map[choice]
        }

        try:
            r = requests.post(f"{API_URL}/admin/monitoring/set_absence", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех", f"Статус установлен на период!\n{r.json().get('period', '')}")
                self.load_monitoring_status()
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))


    def take_raw_material_dialog(self):
        if self.raw_table.currentRow() < 0:
            QMessageBox.warning(self, "Внимание", "Выберите материал из таблицы!")
            return

        row = self.raw_table.currentRow()
        mat_id = self.raw_table.item(row, 0).text()
        mat_name = self.raw_table.item(row, 1).text()

        qty, ok = QInputDialog.getDouble(self, f"Взять: {mat_name}", "Сколько взять?", 1.0, decimals=1)
        if not ok or qty <= 0: return

        try:
            payload = {"material_id": int(mat_id), "amount": qty}
            r = requests.post(f"{API_URL}/admin/warehouse/raw/take", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех", f"Материал списан! Остаток: {r.json()['remaining']}")
                self.load_raw_materials()
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка сервера"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def add_raw_material_dialog(self):
        """Диалог добавления сырья на склад"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить сырье на склад")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Название
        name_input = QLineEdit()
        name_input.setPlaceholderText("Например: Полипропилен")
        layout.addRow("Название:", name_input)

        # Толщина/Размер
        thickness_input = QLineEdit()
        thickness_input.setPlaceholderText("Например: 3 мм")
        layout.addRow("Толщина (размер):", thickness_input)

        # Цвет
        color_input = QLineEdit()
        color_input.setPlaceholderText("Например: прозрачный")
        layout.addRow("Цвет:", color_input)

        # Количество
        qty_input = QLineEdit()
        qty_input.setPlaceholderText("0")
        layout.addRow("Количество (шт):", qty_input)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_input.text().strip().lower()
            thickness = thickness_input.text().strip()
            # Нормализуем толщину - заменяем точку на запятую и добавляем "мм" если нет
            thickness = thickness.replace('.', ',').strip()
            if not thickness.lower().endswith('мм'):
                thickness = thickness + 'мм'

            color = color_input.text().strip().lower()
            qty_text = qty_input.text().strip()

            if not name or not thickness or not color or not qty_text:
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                qty = float(qty_text)
                if qty <= 0:
                    raise ValueError

                payload = {
                    "name": name,
                    "thickness": thickness,
                    "color": color,
                    "quantity": qty
                }

                r = requests.post(
                    f"{API_URL}/admin/warehouse/raw/add",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 201:
                    QMessageBox.information(self, "Успех", f"Сырье '{name}' добавлено!\nКоличество: {qty} шт.")
                    self.load_raw_materials()
                else:
                    error_detail = r.json().get("detail", "Ошибка добавления")
                    if isinstance(error_detail, list):
                        error_detail = " | ".join(str(item) for item in error_detail)
                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Неверное количество!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def _title_label(self, text):
        lbl = QLabel(text)
        lbl.setFont(QFont('Segoe UI', 28, QFont.Weight.Bold))
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        return lbl

    def _toggle_secret(self, text):
        is_admin = text == "Администратор"
        is_manager = text == "Руководитель"
        self.secret_widget.setVisible(is_admin or is_manager)

        if is_manager:
            self.reg_secret.setPlaceholderText("BOSS_MANAGER_KEY_2026")
        else:
            self.reg_secret.setPlaceholderText("SUPER_SECRET_ADMIN_KEY_2026")

    def login(self):
        u, p = self.login_input.text().strip(), self.password_input.text()
        if not u or not p:
            return QMessageBox.warning(self, "Ошибка", "Введите логин и пароль")

        try:
            r = requests.post(f"{API_URL}/auth/login", data={"username": u, "password": p})
            if r.status_code == 200:
                d = r.json()
                self.load_production_data()
                self.token = d["access_token"]
                self.user_role = d["role"]
                self.username = u
                self.load_dashboard()

                # === 🔥 ПОЛНОЕ ПЕРЕСОЗДАНИЕ ВКЛАДКИ "СМЕНЫ" 🔥 ===
                # 1. Удаляем старую вкладку "Смены" (индекс 0)
                if self.tabs.count() > 0:
                    self.tabs.removeTab(0)

                # 2. Создаем НОВЫЙ виджет с нуля
                self.shifts_tab = QWidget()

                # 3. Вставляем новую вкладку на первое место
                self.tabs.insertTab(0, self.shifts_tab, "Смены")

                # 4. Пересобираем интерфейс под текущую роль
                self._setup_shifts_interface()

                # 5. Возвращаемся на вкладку "Смены"
                self.tabs.setCurrentIndex(0)
                # ===================================================

                # --- ОЧИСТКА ЛИШНИХ ВКЛАДОК ---
                while self.tabs.count() > 2:
                    self.tabs.removeTab(2)

                # --- ДОБАВЛЕНИЕ СПЕЦИФИЧНЫХ ВКЛАДОК ---
                if self.user_role == "admin":
                    # Пересоздаем вкладку "Склад сырья" для админа
                    self.load_production_data()
                    self.raw_tab = QWidget()
                    self.setup_raw_tab()
                    self.tabs.addTab(self.raw_tab, "Склад сырья")
                    self.tabs.addTab(self.monitoring_tab, "Мониторинг")
                    self.tabs.addTab(self.blanks_tab, "Склад заготовок")
                    self.load_monitoring_status()
                    self.load_raw_materials()
                    self.load_blanks_list()
                    self.vacation_tab = QWidget()
                    self.setup_vacation_tab()
                    self.tabs.addTab(self.vacation_tab, "График отпусков")
                    self.tabs.addTab(self.recipes_tab, "Рецептуры")

                    self.load_recipes()
                    self.load_monitoring_status()
                    self.load_raw_materials()
                    self.load_blanks_list()
                    self.load_vacation_schedule()


                elif self.user_role == "manager":
                    # Пересоздаем вкладку "Склад сырья" для руководителя

                    self.load_production_data()
                    self.raw_tab = QWidget()
                    self.setup_raw_tab()
                    self.tabs.addTab(self.raw_tab, "Склад сырья")

                    # Добавляем вкладку "Склад заготовок"
                    self.blanks_tab = QWidget()
                    self.setup_blanks_tab()
                    self.tabs.addTab(self.blanks_tab, "Склад заготовок")
                    self.tabs.addTab(self.finance_tab, "Финансы")
                    self.tabs.addTab(self.payroll_tab, "Зарплата")

                    # Добавляем вкладку "Рецептуры"
                    self.recipes_tab = QWidget()
                    self.setup_recipes_tab()
                    self.tabs.addTab(self.recipes_tab, "Рецептуры")
                    self.tabs.addTab(self.vacation_tab, "График отпусков")
                    self.load_finance_data()
                    self.load_raw_materials()
                    self.load_blanks_list()
                    self.load_recipes()
                    self.load_vacation_schedule()
                self.stack.setCurrentIndex(2)
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка входа"))
        except requests.exceptions.ConnectionError:
            QMessageBox.critical(self, "Ошибка", "Нет связи с сервером.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def register(self):
        login, pwd = self.reg_login.text().strip(), self.reg_pass.text()
        if not login or not pwd:
            return QMessageBox.warning(self, "Ошибка", "Заполните логин и пароль")

        role_text = self.reg_role.currentText()

        # Правильная маппинг ролей
        if role_text == "Администратор":
            role = "admin"
        elif role_text == "Руководитель":
            role = "manager"
        else:
            role = "user"

        payload = {"login": login, "password": pwd, "role": role}

        # Секретный ключ нужен и админу, и руководителю
        if role in ["admin", "manager"]:
            payload["admin_secret"] = self.reg_secret.text().strip()

        try:
            r = requests.post(f"{API_URL}/auth/register", json=payload)
            if r.status_code == 201:
                QMessageBox.information(self, "Успех", "Аккаунт создан! Войдите.")
                self.stack.setCurrentIndex(0)
                [w.clear() for w in (self.reg_login, self.reg_pass, self.reg_secret)]
                self.reg_role.setCurrentIndex(0)
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def load_dashboard(self):
        self.welcome.setText(f"Привет, {self.username}")
        self.load_shifts_stats()
        self.load_production_data()
        if self.user_role == "admin":
            self.load_blanks_list()

    def load_shifts_stats(self):
        try:
            r = requests.get(f"{API_URL}/shifts/history?limit=1000", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                shifts = r.json()
                self.l_shifts.setText(str(len(shifts)))
                self.l_hours.setText(f"{sum(s.get('duration_hours', 0) or 0 for s in shifts):.1f}")
                has_open = any(not s.get("end_time") for s in shifts)
                self.open_btn.setVisible(not has_open)
                self.close_btn.setVisible(has_open)
        except:
            pass

    def open_shift(self):
        try:
            r = requests.post(f"{API_URL}/shifts/open", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Смена открыта", "Успешно!")
                self.load_dashboard()
            else:
                QMessageBox.warning(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def close_shift(self):
        try:
            r = requests.post(f"{API_URL}/shifts/close", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Смена закрыта", f"Отработано: {r.json().get('duration_hours', 0)} ч.")
                self.load_dashboard()
            else:
                QMessageBox.warning(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def show_history(self):
        try:
            r = requests.get(f"{API_URL}/shifts/history", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                shifts = r.json()
                if not shifts: return QMessageBox.information(self, "История", "Пусто")
                msg = "История смен:\n"
                for i, s in enumerate(shifts[:10], 1):
                    st = datetime.fromisoformat(s["start_time"].replace("Z", "+00:00"))
                    line = f"{i}. {st.strftime('%d.%m %H:%M')}"
                    if s.get("end_time"):
                        et = datetime.fromisoformat(s["end_time"].replace("Z", "+00:00"))
                        line += f" - {et.strftime('%H:%M')} ({s['duration_hours']}ч)"
                    if s.get("is_late"): line += f" Опоздание: +{s['late_minutes']}мин"
                    msg += line + "\n"
                QMessageBox.information(self, "История", msg)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def show_all_shifts_monitoring(self):
        """Показывает смены всех сотрудников с текущим статусом (Менеджер)"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Мониторинг смен - Все сотрудники")
        dialog.setMinimumSize(800, 600)

        layout = QVBoxLayout(dialog)

        # Заголовок
        title = QLabel("<h2> Мониторинг сотрудников</h2>")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Кнопка обновления
        refresh_btn = QPushButton("🔄 Обновить статусы")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background: #007bff;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        layout.addWidget(refresh_btn)

        # Таблица
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Сотрудник", "Статус", "Время начала", "Длительность"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(table)

        # Легенда
        legend = QLabel("<b>Легенда:</b> 🟢 Работает | 🔵 Опоздал | 🔴 Не вышел | 🟡 Болен | 🟣 Отпуск")
        legend.setStyleSheet("color: #555; font-size: 11px; padding: 5px;")
        layout.addWidget(legend)

        # Внутренняя функция загрузки
        def load_data():
            try:
                # 1. Делаем запрос к НОВОМУ менеджерскому эндпоинту
                url = f"{API_URL}/manager/monitoring/status?target_date={datetime.now().strftime('%Y-%m-%d')}"
                r = requests.get(url, headers={"Authorization": f"Bearer {self.token}"})

                if r.status_code == 200:
                    employees = r.json()
                    table.setRowCount(len(employees))

                    for i, emp in enumerate(employees):
                        # Логин
                        table.setItem(i, 0, QTableWidgetItem(emp['login']))

                        # Статус с цветом
                        status_text = self._get_status_text(emp['status'])
                        status_item = QTableWidgetItem(status_text)
                        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        status_item.setBackground(self._get_status_color(emp['status']))
                        table.setItem(i, 2, status_item)

                        # Время и длительность
                        table.setItem(i, 3, QTableWidgetItem(emp.get('start_time') or '-'))
                        table.setItem(i, 4, QTableWidgetItem(emp.get('duration') or '-'))

                    table.resizeRowsToContents()
                else:
                    # Если ошибка (например, 403 или 500)
                    QMessageBox.warning(self, "Ошибка", f"Ошибка сервера: {r.status_code}\n{r.text}")

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить: {str(e)}")

        refresh_btn.clicked.connect(load_data)
        load_data()  # Загружаем сразу при открытии

        dialog.exec()

    # --- ПРОИЗВОДСТВО ---

    def load_production_data(self):
        """Загрузка заготовок и истории производства"""
        print("🔄 Загрузка данных производства...")

        # === ЧАСТЬ 1: ЗАГРУЗКА ЗАГОТОВОК ===
        try:
            r = requests.get(f"{API_URL}/warehouse/blanks", headers={"Authorization": f"Bearer {self.token}"})
            print(f"📡 Запрос заготовок: статус {r.status_code}")

            if r.status_code == 200:
                blanks = r.json()
                print(f"✅ Получено заготовок: {len(blanks)}")

                self.blank_combo.clear()
                self.blank_ids.clear()

                if not blanks:
                    self.blank_combo.addItem("Нет заготовок")
                    print("⚠️ Список заготовок пуст!")
                else:
                    for b in blanks:
                        display_text = f"{b['name']} (Остаток: {b['quantity']})"
                        self.blank_combo.addItem(display_text, userData=b['id'])
                        self.blank_ids.append(b['id'])
                        print(f"  ➕ {display_text} (ID: {b['id']})")

                    print(f"✅ Всего элементов в combo: {self.blank_combo.count()}")
            else:
                print(f"❌ Ошибка сервера: {r.status_code} - {r.text}")

        except requests.exceptions.ConnectionError:
            print("❌ НЕТ СВЯЗИ С СЕРВЕРОМ!")
        except Exception as e:
            print(f"❌ Ошибка загрузки заготовок: {e}")
            import traceback
            traceback.print_exc()

        # === ЧАСТЬ 2: ЗАГРУЗКА ИСТОРИИ ===
        try:
            r = requests.get(f"{API_URL}/user/production/history",
                             headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                logs = r.json()
                self.prod_table.setRowCount(len(logs))
                for i, log in enumerate(logs):
                    date_str = datetime.fromisoformat(log['created_at'].replace("Z", "+00:00")).strftime("%d.%m %H:%M")
                    self.prod_table.setItem(i, 0, QTableWidgetItem(date_str))
                    self.prod_table.setItem(i, 1, QTableWidgetItem(log.get('product_name', '-')))
                    self.prod_table.setItem(i, 2, QTableWidgetItem(str(log['blanks_taken'])))
                    self.prod_table.setItem(i, 3, QTableWidgetItem(str(log['items_produced'])))
                    self.prod_table.setItem(i, 4, QTableWidgetItem(str(log['defect_amount'])))
                print(f"✅ Загружено записей истории: {len(logs)}")
            else:
                print(f"❌ Ошибка загрузки истории: {r.status_code}")
        except Exception as e:
            print(f"❌ Ошибка загрузки истории: {e}")

    def submit_production_report(self):
        """Отправка отчёта о производстве со списанием заготовок"""

        # Проверка данных
        if self.blank_combo.count() == 0 or self.blank_combo.currentText() in ["Нет заготовок",
                                                                               "Загрузка заготовок..."]:
            QMessageBox.warning(self, "Ошибка", "Выберите заготовку из списка!")
            return

        # Получаем ID заготовки
        blank_id = self.blank_combo.currentData()

        # Если currentData() вернул None - берём из списка blank_ids
        if blank_id is None:
            current_index = self.blank_combo.currentIndex()
            if hasattr(self, 'blank_ids') and 0 <= current_index < len(self.blank_ids):
                blank_id = self.blank_ids[current_index]
                print(f"✅ ID получен из blank_ids: {blank_id}")
            else:
                QMessageBox.warning(self, "Ошибка", "Не выбрано изделие! Выберите заготовку из списка.")
                return

        if blank_id is None:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить ID изделия!")
            return

        try:
            blanks_taken = int(self.in_taken.text() or 0)
            items_produced = int(self.in_produced.text() or 0)
            defect_amount = int(self.in_defect.text() or 0)
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Количество должно быть числом!")
            return

        reason = self.in_reason.text().strip()

        if blanks_taken <= 0:
            QMessageBox.warning(self, "Ошибка", "Укажите количество взятых заготовок")
            return

        # Логика недостачи
        total_output = items_produced + defect_amount
        shortage = blanks_taken - total_output

        # Если есть недостача, требуем причину
        if shortage > 0:
            reason, ok = QInputDialog.getText(
                self, "ВНИМАНИЕ: Недостача",
                f"Вы взяли {blanks_taken}, а сдали {total_output}.\nНедостача: {shortage} шт.\nУкажите причину:"
            )
            if not ok or not reason.strip():
                QMessageBox.warning(self, "Ошибка", "Необходимо указать причину недостачи!")
                return
            reason = reason.strip()

            # ВАЖНО: добавляем недостачу к браку
            defect_amount += shortage
            print(f"⚠️ Недостача {shortage} шт. добавлена к браку. Итого брак: {defect_amount}")

        # Получаем название заготовки для product_name
        selected_blank_name = self.blank_combo.currentText()
        if " (Остаток:" in selected_blank_name:
            product_name = selected_blank_name.split(" (Остаток:")[0].strip()
        else:
            product_name = selected_blank_name

        # === ОТПРАВКА ОТЧЁТА И СПИСАНИЕ ЗАГОТОВОК ===
        try:
            # 1. Сначала отправляем отчёт о производстве
            payload = {
                "blank_id": int(blank_id),
                "blanks_taken": blanks_taken,
                "items_produced": items_produced,
                "defect_amount": defect_amount,  # Теперь включает недостачу
                "defect_reason": reason if shortage > 0 else None,
                "product_name": product_name
            }

            print(f"📤 Отправляем отчёт: {payload}")

            r = requests.post(
                f"{API_URL}/user/production/report",
                json=payload,
                headers={"Authorization": f"Bearer {self.token}"}
            )

            if r.status_code != 200:
                try:
                    error_detail = r.json().get("detail", "Неизвестная ошибка")
                    if isinstance(error_detail, list):
                        error_detail = " | ".join(str(item) for item in error_detail)
                    QMessageBox.critical(self, "Ошибка", str(error_detail))
                except:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка сервера: {r.status_code}")
                return

            # 2. ✅ СПИСЫВАЕМ заготовки со склада
            print(f"📦 Списываем {blanks_taken} заготовок со склада...")
            take_payload = {
                "blank_id": int(blank_id),
                "amount": blanks_taken
            }

            take_r = requests.post(
                f"{API_URL}/warehouse/blanks/take",
                json=take_payload,
                headers={"Authorization": f"Bearer {self.token}"}
            )

            if take_r.status_code == 200:
                remaining = take_r.json().get('remaining', blanks_taken)
                print(f"✅ Заготовки списаны! Остаток: {remaining}")
            else:
                print(f"⚠️ Не удалось списать заготовки: {take_r.status_code}")

            # 3. Успех!
            QMessageBox.information(self, "Успех", f"Отчёт принят!\nЗаготовки списаны со склада.")

            # Очистка полей
            self.in_taken.setText("1")
            self.in_produced.setText("0")
            self.in_defect.setText("0")
            self.in_reason.clear()

            # Обновляем данные (заготовки и историю)
            self.load_production_data()

        except requests.exceptions.ConnectionError:
            QMessageBox.critical(self, "Ошибка", "Нет связи с сервером. Убедитесь, что сервер запущен.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось отправить отчёт: {str(e)}")

    # --- СКЛАД ЗАГОТОВОК (АДМИН) ---

    def load_blanks_list(self):
        try:
            r = requests.get(f"{API_URL}/warehouse/blanks", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                blanks = r.json()
                self.blanks_table.setRowCount(len(blanks))
                for i, b in enumerate(blanks):
                    print(f"  [{i}] Заготовка: ID={b['id']}, Name={b['name']}, Qty={b['quantity']}")

                    self.blanks_table.setItem(i, 0, QTableWidgetItem(str(b['id'])))
                    self.blanks_table.setItem(i, 1, QTableWidgetItem(b['name']))
                    self.blanks_table.setItem(i, 2, QTableWidgetItem(str(b['quantity'])))

                    # Пустая ячейка (колонка 3)
                    self.blanks_table.setItem(i, 3, QTableWidgetItem(""))

                    # Контейнер для кнопок действий (колонка 4)
                    container = QWidget()
                    layout = QHBoxLayout(container)
                    layout.setContentsMargins(2, 2, 2, 2)
                    layout.setSpacing(2)

                    # Кнопка Добавить (новая)
                    add_btn = QPushButton("➕")
                    add_btn.setToolTip("Добавить количество")
                    add_btn.setFixedSize(30, 30)
                    add_btn.setStyleSheet("""
                        QPushButton {
                            background: #17a2b8;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                        }
                        QPushButton:hover {
                            background: #138496;
                        }
                    """)
                    add_btn.clicked.connect(
                        lambda checked, bid=b['id'], bname=b['name'], bqty=b['quantity']:
                        self.add_quantity_dialog(bid, bname, bqty)
                    )

                    # Кнопка Изменить
                    edit_btn = QPushButton("✏️")
                    edit_btn.setToolTip("Изменить заготовку")
                    edit_btn.setFixedSize(30, 30)
                    edit_btn.setStyleSheet("""
                        QPushButton {
                            background: #28a745;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                        }
                        QPushButton:hover {
                            background: #218838;
                        }
                    """)
                    edit_btn.clicked.connect(
                        lambda checked, bid=b['id'], bname=b['name'], bqty=b['quantity']:
                        self.edit_blank_from_table(bid, bname, bqty)
                    )

                    # Кнопка Удалить
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setToolTip("Удалить заготовку")
                    delete_btn.setFixedSize(30, 30)
                    delete_btn.setStyleSheet("""
                        QPushButton {
                            background: #dc3545;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                        }
                        QPushButton:hover {
                            background: #c82333;
                        }
                    """)
                    delete_btn.clicked.connect(
                        lambda checked, bid=b['id'], bname=b['name']:
                        self.delete_blank(bid, bname)
                    )

                    layout.addWidget(add_btn)
                    layout.addWidget(edit_btn)
                    layout.addWidget(delete_btn)

                    self.blanks_table.setCellWidget(i, 4, container)

                    # Выравнивание по центру
                    self.blanks_table.item(i, 2).setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                # Скрываем колонку с ID
                self.blanks_table.setColumnHidden(0, True)
                # Скрываем пустую колонку
                self.blanks_table.setColumnHidden(3, True)
                # Обновляем заголовки
                self.blanks_table.setHorizontalHeaderLabels(["ID", "Название", "Остаток", "", "Действия"])

                print(f"✅ Загружено заготовок: {len(blanks)}")
            else:
                print(f"❌ Ошибка загрузки заготовок: {r.status_code}")
        except Exception as e:
            print(f"Ошибка загрузки заготовок: {e}")
            import traceback
            traceback.print_exc()

    def add_quantity_dialog(self, blank_id, blank_name, current_qty):
        """Диалог быстрого добавления количества"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Добавить: {blank_name}")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Текущее количество
        current_label = QLabel(f"{current_qty} шт.")
        current_label.setStyleSheet("font-weight: bold; color: #ffc107; font-size: 14px;")
        layout.addRow("Текущее количество:", current_label)

        # Сколько добавить
        add_input = QLineEdit()
        add_input.setPlaceholderText("Сколько добавить?")
        layout.addRow("Добавить (шт):", add_input)

        # Итоговое количество (информация)
        self.temp_total_label = QLabel(f"Будет: {current_qty} шт.")
        self.temp_total_label.setStyleSheet("color: #28a745; font-weight: bold;")
        layout.addRow("После добавления:", self.temp_total_label)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        # Обновляем итог при изменении текста
        def update_total(text):
            try:
                add_qty = int(text) if text else 0
                total = current_qty + add_qty
                self.temp_total_label.setText(f"Будет: {total} шт.")
            except:
                self.temp_total_label.setText(f"Будет: {current_qty} шт.")

        add_input.textChanged.connect(update_total)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            add_text = add_input.text().strip()

            if not add_text:
                QMessageBox.warning(self, "Ошибка", "Введите количество!")
                return

            try:
                add_qty = int(add_text)
                if add_qty <= 0:
                    raise ValueError

                payload = {
                    "name": blank_name,
                    "quantity": add_qty
                }

                r = requests.post(f"{API_URL}/admin/warehouse/blanks/add",
                                  json=payload,
                                  headers={"Authorization": f"Bearer {self.token}"})

                if r.status_code == 201:
                    new_total = current_qty + add_qty
                    QMessageBox.information(
                        self, "Успех",
                        f"Добавлено {add_qty} шт.!\n"
                        f"Всего: {new_total} шт."
                    )
                    self.load_blanks_list()
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Введите положительное число!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def add_blank_dialog(self):
        """Диалог добавления заготовок"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить заготовки")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Название
        name_input = QLineEdit()
        name_input.setPlaceholderText("Например: Бокс №10")
        layout.addRow("Название:", name_input)

        # Количество
        qty_input = QLineEdit()
        qty_input.setPlaceholderText("0")
        layout.addRow("Количество (шт):", qty_input)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_input.text().strip()
            qty_text = qty_input.text().strip()

            if not name or not qty_text:
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                qty = int(qty_text)
                if qty <= 0:
                    raise ValueError

                payload = {"name": name, "quantity": qty}
                r = requests.post(f"{API_URL}/admin/warehouse/blanks/add", json=payload,
                                  headers={"Authorization": f"Bearer {self.token}"})
                if r.status_code == 201:
                    QMessageBox.information(self, "Успех", f"Заготовки '{name}' добавлены!\nКоличество: {qty} шт.")
                    self.load_blanks_list()
                    self.load_production_data()  # Обновить выпадающий список в производстве
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Неверное количество!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def take_blank_dialog(self):
        if self.blanks_table.currentRow() < 0:
            QMessageBox.warning(self, "Внимание", "Выберите заготовку из таблицы!")
            return

        row = self.blanks_table.currentRow()
        blank_id = self.blanks_table.item(row, 0).text()
        blank_name = self.blanks_table.item(row, 1).text()
        current_qty = int(self.blanks_table.item(row, 2).text())

        # Исправлено: minValue и maxValue вместо min и max
        amount, ok = QInputDialog.getInt(
            self,
            f"Взять: {blank_name}",
            f"Доступно: {current_qty} шт.\nСколько взять?",
            1,  # значение по умолчанию
            1,  # minValue (было min=1)
            current_qty,  # maxValue (было max=current_qty)
            1  # step
        )
        if not ok or amount <= 0:
            return

        try:
            payload = {"blank_id": int(blank_id), "amount": amount}
            r = requests.post(f"{API_URL}/warehouse/blanks/take", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех", f"Заготовки взяты! Остаток: {r.json()['remaining']}")
                self.load_blanks_list()
                self.load_production_data()
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def take_blank_from_table(self, blank_id, blank_name, current_qty):
        """Взять заготовки из таблицы (быстрый доступ)"""
        amount, ok = QInputDialog.getInt(
            self,
            f"Взять: {blank_name}",
            f"Доступно: {current_qty} шт.\nСколько взять?",
            1,  # значение по умолчанию
            1,  # minValue
            current_qty,  # maxValue
            1  # step
        )
        if not ok or amount <= 0:
            return

        try:
            payload = {"blank_id": int(blank_id), "amount": amount}
            r = requests.post(f"{API_URL}/warehouse/blanks/take", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех", f"Заготовки взяты! Остаток: {r.json()['remaining']}")
                self.load_blanks_list()
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def edit_blank_from_table(self, blank_id, blank_name, current_qty):
        """Редактирование заготовки (название и/или количество)"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Изменить: {blank_name}")
        dialog.setMinimumWidth(400)

        layout = QFormLayout(dialog)

        # Название
        name_input = QLineEdit(blank_name)
        name_input.setPlaceholderText("Название заготовки")
        layout.addRow("Название:", name_input)

        # Количество
        qty_input = QLineEdit(str(current_qty))
        qty_input.setPlaceholderText("Количество")
        layout.addRow("Количество:", qty_input)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_name = name_input.text().strip()
            qty_text = qty_input.text().strip()

            if not new_name or not qty_text:
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                new_qty = int(qty_text)
                if new_qty < 0:
                    raise ValueError

                payload = {
                    "name": new_name,
                    "quantity": new_qty
                }

                r = requests.put(
                    f"{API_URL}/admin/warehouse/blanks/{blank_id}",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Заготовка '{new_name}' обновлена!")
                    self.load_blanks_list()
                else:
                    QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка обновления"))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Неверное количество!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def delete_blank(self, blank_id, blank_name):
        """Удаление заготовки со склада"""
        reply = QMessageBox.question(
            self,
            "Удаление заготовки",
            f"Вы уверены, что хотите удалить '{blank_name}'?\nЭто действие необратимо!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                r = requests.delete(
                    f"{API_URL}/admin/warehouse/blanks/{blank_id}",
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    # Проверяем что ответ не пустой
                    if r.text.strip():
                        result = r.json()
                        QMessageBox.information(self, "Успех", f"Заготовка '{blank_name}' удалена!")
                    else:
                        QMessageBox.information(self, "Успех", f"Заготовка '{blank_name}' удалена!")

                    self.load_blanks_list()
                else:
                    # Пробуем получить текст ошибки
                    try:
                        error_detail = r.json().get("detail", "Ошибка удаления")
                    except:
                        error_detail = r.text or f"Ошибка сервера: {r.status_code}"

                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def logout(self):
        self.token = self.user_role = self.username = None
        self.login_input.clear()
        self.password_input.clear()
        self.stack.setCurrentIndex(0)
        # Сброс табов
        if self.tabs.count() == 3:
            self.tabs.removeTab(2)

    def closeEvent(self, event):
        if self.server_process:
            try:
                self.server_process.terminate()
                self.server_process.wait(timeout=3)
            except:
                pass
        event.accept()

# --- ФИНАНСЫ ---

    def setup_finance_tab(self):
        layout = QVBoxLayout(self.finance_tab)
        layout.setSpacing(15)

        layout.addWidget(QLabel("<b>Управление способами оплаты:</b>"))

        self.finance_table = QTableWidget()
        self.finance_table.setColumnCount(6)  # <-- ВАЖНО: 6 колонок!
        self.finance_table.setHorizontalHeaderLabels([
            "ID",
            "Сотрудник",
            "Роль",
            "Система оплаты",
            "Ставка (руб)",
            "Действия"
        ])
        self.finance_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.finance_table)

        refresh_btn = QPushButton("🔄 Обновить список")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px;")
        refresh_btn.clicked.connect(self.load_finance_data)
        layout.addWidget(refresh_btn)

    def update_payment_simple(self, text, combo):
        user_id = combo.property("user_id")
        try:
            payload = {"user_id": user_id, "payment_method": text}
            r = requests.post(f"{API_URL}/manager/users/payment", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code != 200:
                QMessageBox.warning(self, "Ошибка", "Не удалось сохранить")
                self.load_finance_data()
        except Exception as e:
            print(f"Ошибка: {e}")
            self.load_finance_data()

    def load_finance_data(self):
        try:
            r = requests.get(f"{API_URL}/manager/users/list", headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                users = r.json()
                self.finance_table.setRowCount(len(users))

                for i, u in enumerate(users):
                    self.finance_table.setItem(i, 0, QTableWidgetItem(str(u['id'])))
                    self.finance_table.setItem(i, 1, QTableWidgetItem(u['login']))
                    self.finance_table.setItem(i, 2, QTableWidgetItem(u['role']))

                    # ComboBox с системой оплаты (колонка 3)
                    combo = QComboBox()
                    combo.addItems(["Не указан", "Почасовая", "Сдельная", "Оклад + премия", "Процент от выработки"])
                    combo.setCurrentText(u['payment_method'])
                    combo.setProperty("user_id", u['id'])
                    combo.currentTextChanged.connect(
                        lambda text, combo=combo: self.on_payment_method_changed(text, combo))
                    self.finance_table.setCellWidget(i, 3, combo)

                    # Ставка (колонка 4)
                    rate_text = f"{u['payment_rate']:.2f}" if u['payment_rate'] else "-"
                    rate_item = QTableWidgetItem(rate_text)
                    rate_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.finance_table.setItem(i, 4, rate_item)  # <-- 5-я колонка (индекс 4)

                    # --- КОЛОНКА ДЕЙСТВИЯ (Кнопки) ---
                    actions_widget = QWidget()
                    actions_layout = QHBoxLayout(actions_widget)
                    actions_layout.setContentsMargins(2, 2, 2, 2)
                    actions_layout.setSpacing(5)

                    # Кнопка Сброс пароля
                    btn_reset = QPushButton("🔑 Сброс")  # Текст явно
                    btn_reset.setToolTip("Сменить пароль сотруднику")
                    btn_reset.setMinimumWidth(80)  # Фиксируем минимальную ширину
                    btn_reset.setStyleSheet("background-color: #007bff; color: white; border-radius: 5px;")
                    btn_reset.clicked.connect(
                        lambda checked, uid=u['id'], login=u['login']: self.reset_password_dialog(uid, login))
                    actions_layout.addWidget(btn_reset)

                    # Кнопка Удалить
                    btn_delete = QPushButton("🗑️ Удалить")  # Текст явно
                    btn_delete.setToolTip("Удалить сотрудника")
                    btn_delete.setMinimumWidth(80)  # Фиксируем минимальную ширину
                    btn_delete.setStyleSheet("background-color: #dc3545; color: white; border-radius: 5px;")
                    btn_delete.clicked.connect(
                        lambda checked, uid=u['id'], login=u['login']: self.delete_user_dialog(uid, login))
                    actions_layout.addWidget(btn_delete)

                    self.finance_table.setCellWidget(i, 5, actions_widget)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def on_payment_method_changed(self, payment_method, combo):
        """Обработка изменения системы оплаты"""
        user_id = combo.property("user_id")

        # Если выбрали "Не указан" - просто сохраняем без ставки
        if payment_method == "Не указан":
            self.save_payment_rate(user_id, payment_method, 0.0)
            return

        # Для остальных систем - запрашиваем ставку
        rate, ok = QInputDialog.getDouble(
            self,
            f"Ставка для {payment_method}",
            f"Введите ставку (в рублях):\n\n"
            f"{'Почасовая оплата (руб/час)' if payment_method == 'Почасовая' else ''}"
            f"{'Сдельная оплата (руб/шт)' if payment_method == 'Сдельная' else ''}"
            f"{'Оклад (руб/мес)' if payment_method == 'Оклад + премия' else ''}"
            f"{'Процент (%)' if payment_method == 'Процент от выработки' else ''}",
            0.0,
            0.0,
            1000000.0,
            2
        )

        if ok and rate > 0:
            self.save_payment_rate(user_id, payment_method, rate)
        else:
            # Если отменили - перезагружаем данные
            self.load_finance_data()

    def save_payment_rate(self, user_id, payment_method, payment_rate):
        """Сохранение системы оплаты и ставки"""
        try:
            payload = {
                "user_id": int(user_id),
                "payment_method": payment_method,
                "payment_rate": float(payment_rate)
            }
            r = requests.post(f"{API_URL}/manager/users/payment", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                QMessageBox.information(self, "Успех",
                                        f"Система оплаты установлена!\n{payment_method}: {payment_rate:.2f} руб.")
                self.load_finance_data()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось сохранить")
                self.load_finance_data()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")
            self.load_finance_data()


    def update_payment(self, text, combo):
        user_id = combo.property("user_id")
        try:
            payload = {"user_id": user_id, "payment_method": text}
            r = requests.post(f"{API_URL}/manager/users/payment", json=payload,
                              headers={"Authorization": f"Bearer {self.token}"})
            if r.status_code == 200:
                print(f"Оплата для {user_id} обновлена на {text}")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось сохранить")
                self.load_finance_data()  # Откатить если ошибка
        except Exception as e:
            print(e)

# ---- ЗАРПЛАТА ----

    def setup_payroll_tab(self):
        layout = QVBoxLayout(self.payroll_tab)
        layout.setSpacing(15)

        # Панель выбора периода
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Период расчета:"))

        self.payroll_start = QLineEdit(datetime.now().replace(day=1).strftime("%Y-%m-%d"))
        self.payroll_start.setMaximumWidth(120)
        period_layout.addWidget(self.payroll_start)

        period_layout.addWidget(QLabel("—"))

        self.payroll_end = QLineEdit(datetime.now().strftime("%Y-%m-%d"))
        self.payroll_end.setMaximumWidth(120)
        period_layout.addWidget(self.payroll_end)

        calc_btn = QPushButton("Рассчитать")
        calc_btn.setStyleSheet(
            "background: #fd7e14; color: white; border-radius: 8px; "
            "font-weight: bold; font-size: 14px;"
        )
        calc_btn.setMinimumHeight(36)
        calc_btn.clicked.connect(self.calculate_payroll)
        period_layout.addWidget(calc_btn)

        period_layout.addStretch()
        layout.addLayout(period_layout)

        # Таблица с результатами
        self.payroll_table = QTableWidget()
        self.payroll_table.setColumnCount(3)  # ID, Имя, Итого (дни добавим динамически)
        self.payroll_table.setHorizontalHeaderLabels(["Сотрудник", "Система", "Итого (руб)"])
        self.payroll_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.payroll_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.payroll_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.payroll_table)

        # Кнопка экспорта
        export_btn = QPushButton("Экспорт в отчет")
        export_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px;")
        export_btn.clicked.connect(self.export_payroll)
        layout.addWidget(export_btn)

    def calculate_payroll(self):
        start_date = self.payroll_start.text()
        end_date = self.payroll_end.text()

        try:
            r = requests.get(
                f"{API_URL}/manager/payroll/calculate?start_date={start_date}&end_date={end_date}",
                headers={"Authorization": f"Bearer {self.token}"}
            )
            if r.status_code == 200:
                data = r.json()
                self.payroll_data = data
                self._display_payroll(data)
            else:
                QMessageBox.critical(self, "Ошибка", r.json().get("detail", "Ошибка расчета"))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось рассчитать: {str(e)}")

    def _display_payroll(self, data):
        employees = data.get("employees", [])
        if not employees:
            QMessageBox.information(self, "Инфо", "Нет данных для расчета за выбранный период")
            return

        # Собираем все уникальные даты для колонок
        all_dates = set()
        for emp in employees:
            all_dates.update(emp["daily"].keys())
        all_dates = sorted(all_dates)

        # Настраиваем таблицу: Сотрудник | Система | Дни... | Итого
        col_count = 3 + len(all_dates)  # Имя, Система, Дни, Итого
        self.payroll_table.setColumnCount(col_count)

        headers = ["Сотрудник", "Система"] + [d[5:] for d in all_dates] + ["Итого (руб)"]
        self.payroll_table.setHorizontalHeaderLabels(headers)

        # Заполняем данными
        self.payroll_table.setRowCount(len(employees))
        for i, emp in enumerate(employees):
            # Имя
            name_item = QTableWidgetItem(emp["login"])
            name_item.setForeground(QColor("#e0e0e0"))  # Светлый текст для тёмной темы
            self.payroll_table.setItem(i, 0, name_item)

            # Система оплаты
            system_item = QTableWidgetItem(f"{emp['payment_method']} ({emp['rate']} руб)")
            system_item.setForeground(QColor("#e0e0e0"))  # Светлый текст
            self.payroll_table.setItem(i, 1, system_item)

            # Дни
            for j, date in enumerate(all_dates, start=2):
                amount = emp["daily"].get(date, 0)
                item = QTableWidgetItem(f"{amount:.2f}" if amount > 0 else "-")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight)

                if amount > 0:
                    item.setBackground(QColor("#28a745"))  # Зелёный фон
                    item.setForeground(QColor("#ffffff"))  # Белый текст (контрастный)
                else:
                    item.setForeground(QColor("#6c757d"))  # Серый текст для "-"

                self.payroll_table.setItem(i, j, item)

            # Итого
            total_item = QTableWidgetItem(f"{emp['total']:.2f}")
            total_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            total_item.setBackground(QColor("#007bff"))
            total_item.setForeground(QColor("white"))
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.payroll_table.setItem(i, col_count - 1, total_item)

    def export_payroll(self):
        """Экспорт расчета зарплаты в Excel"""
        if not hasattr(self, 'payroll_data') or not self.payroll_data.get('employees'):
            QMessageBox.warning(self, "Ошибка", "Сначала рассчитайте зарплату!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет", "Отчет_Зарплата.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path: return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Зарплата"

            # Заголовок с периодом
            period = self.payroll_data['period']
            ws.append([f"Отчет по зарплате: {period['start']} — {period['end']}"])
            ws.append([])  # пустая строка

            employees = self.payroll_data['employees']
            # Собираем все уникальные даты для колонок
            all_dates = sorted(set(d for emp in employees for d in emp['daily'].keys()))

            # Шапка таблицы
            headers = ["Сотрудник", "Система оплаты"] + all_dates + ["Итого (руб)"]
            ws.append(headers)

            # Стилизация шапки (фиолетовый градиент как в приложении)
            header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Заполнение данных
            for emp in employees:
                row = [emp['login'], f"{emp['payment_method']} ({emp['rate']} руб)"]
                for date in all_dates:
                    row.append(emp['daily'].get(date, 0))
                row.append(emp['total'])
                ws.append(row)

            # Стиль колонки "Итого" (синий фон, белый текст)
            total_col = len(headers)
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=total_col)
                cell.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=11)

            # Автоширина колонок
            for col in ws.columns:
                max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 20)

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет сохранен!\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить:\n{str(e)}")

    def setup_recipes_tab(self):
        """Создание вкладки рецептур"""
        layout = QVBoxLayout(self.recipes_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)

        # Заголовок
        layout.addWidget(QLabel("<b>Технологические карты (рецептуры)</b>"))
        layout.addWidget(QLabel("Здесь настраивается сколько изделий помещается в одном листе материала"))

        # Панель управления
        controls = QHBoxLayout()

        add_btn = QPushButton("➕ Добавить рецептуру")
        add_btn.setStyleSheet("""
            QPushButton {
                background: #28a745;
                color: white;
                border-radius: 8px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background: #218838;
            }
        """)
        add_btn.setMinimumHeight(40)
        add_btn.clicked.connect(self.add_recipe_dialog)
        controls.addWidget(add_btn)

        controls.addStretch()
        layout.addLayout(controls)

        # Таблица - 6 колонок (ID скрыт + 4 данных + Действия)
        self.recipes_table = QTableWidget()
        self.recipes_table.setColumnCount(5)
        self.recipes_table.setHorizontalHeaderLabels([
            "ID", "Изделие", "Толщина материала", "Изделий в листе", "Действия", ""
        ])
        self.recipes_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.recipes_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.recipes_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        layout.addWidget(self.recipes_table)

        # Кнопка обновления
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.setStyleSheet("background: #6c757d; color: white; border-radius: 8px; font-weight: bold;")
        refresh_btn.setMinimumHeight(40)
        refresh_btn.clicked.connect(self.load_recipes)
        layout.addWidget(refresh_btn)

    def load_recipes(self):
        """Загрузка списка рецептур"""
        print(f"🔄 Загрузка рецептур...")

        try:
            r = requests.get(f"{API_URL}/recipes/list")
            print(f"📡 Статус ответа: {r.status_code}")

            if r.status_code == 200:
                recipes = r.json()
                print(f"✅ Получено рецептур: {len(recipes)}")

                self.recipes_table.setRowCount(len(recipes))

                for i, recipe in enumerate(recipes):
                    print(f"  [{i}] {recipe['product_name']}: {recipe['material_thickness']}, "
                          f"{recipe['blanks_per_sheet']} шт/лист")

                    self.recipes_table.setItem(i, 0, QTableWidgetItem(str(recipe['id'])))
                    self.recipes_table.setItem(i, 1, QTableWidgetItem(recipe['product_name']))
                    self.recipes_table.setItem(i, 2, QTableWidgetItem(recipe['material_thickness']))
                    self.recipes_table.setItem(i, 3, QTableWidgetItem(str(recipe['blanks_per_sheet'])))

                    # === СОЗДАЁМ КОНТЕЙНЕР С КНОПКАМИ ===
                    container = QWidget()
                    btn_layout = QHBoxLayout(container)
                    btn_layout.setContentsMargins(2, 2, 2, 2)
                    btn_layout.setSpacing(5)

                    # Кнопка Изменить
                    edit_btn = QPushButton("✏️")
                    edit_btn.setToolTip("Изменить рецептуру")
                    edit_btn.setMinimumWidth(40)
                    edit_btn.setStyleSheet("""
                        QPushButton {
                            background: #28a745;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                            padding: 5px;
                        }
                        QPushButton:hover {
                            background: #218838;
                        }
                    """)
                    edit_btn.clicked.connect(
                        lambda checked, rid=recipe['id'], rname=recipe['product_name'],
                               rthickness=recipe['material_thickness'],
                               rblanks=recipe['blanks_per_sheet']:
                        self.edit_recipe_dialog(rid, rname, rthickness, rblanks)
                    )

                    # Кнопка Удалить
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setToolTip("Удалить рецептуру")
                    delete_btn.setMinimumWidth(40)
                    delete_btn.setStyleSheet("""
                        QPushButton {
                            background: #dc3545;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                            padding: 5px;
                        }
                        QPushButton:hover {
                            background: #c82333;
                        }
                    """)
                    delete_btn.clicked.connect(
                        lambda checked, rid=recipe['id'], rname=recipe['product_name']:
                        self.delete_recipe(rid, rname)
                    )

                    btn_layout.addWidget(edit_btn)
                    btn_layout.addWidget(delete_btn)

                    self.recipes_table.setCellWidget(i, 4, container)
                    print(f"  ✅ Кнопки добавлены в строку {i}")

                # Скрываем ID
                self.recipes_table.setColumnHidden(0, True)

                print(f"✅ Загружено рецептур: {len(recipes)}")
            else:
                print(f"❌ Ошибка загрузки: {r.status_code}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить: {r.status_code}")
        except Exception as e:
            print(f"❌ Исключение: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", str(e))

    def edit_recipe_dialog(self, recipe_id, product_name, material_thickness, blanks_per_sheet):
        """Диалог редактирования рецептуры"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Изменить: {product_name}")
        dialog.setMinimumWidth(450)

        layout = QFormLayout(dialog)

        # Название изделия
        name_input = QLineEdit(product_name)
        name_input.setPlaceholderText("Например: Бокс №10")
        layout.addRow("Название изделия:", name_input)

        # Толщина материала
        thickness_combo = QComboBox()
        thickness_combo.addItems(["2мм", "1.5мм", "3мм"])
        thickness_combo.setEditable(True)
        thickness_combo.setCurrentText(material_thickness)
        layout.addRow("Толщина материала:", thickness_combo)

        # Заготовок в одном листе
        blanks_input = QLineEdit(str(blanks_per_sheet))
        blanks_input.setPlaceholderText("Сколько заготовок помещается в 1 листе")
        layout.addRow("Заготовок в 1 листе:", blanks_input)

        # Пояснение
        info_label = QLabel("💡 Пример: в 1 листе 3мм помещается 6 заготовок")
        info_label.setStyleSheet("color: #ffc107; font-style: italic;")
        layout.addRow(info_label)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            product_name_new = name_input.text().strip()
            thickness = thickness_combo.currentText()
            blanks_text = blanks_input.text().strip()

            if not all([product_name_new, thickness, blanks_text]):
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                blanks_per_sheet_new = int(blanks_text)
                if blanks_per_sheet_new <= 0:
                    raise ValueError

                payload = {
                    "product_name": product_name_new.lower(),
                    "material_thickness": thickness,
                    "blanks_per_sheet": blanks_per_sheet_new
                }

                r = requests.put(
                    f"{API_URL}/admin/recipes/{recipe_id}",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    QMessageBox.information(
                        self, "Успех",
                        f"Рецептура обновлена!\n\n"
                        f"Изделие: {product_name_new}\n"
                        f"Толщина: {thickness}\n"
                        f"Заготовок в листе: {blanks_per_sheet_new}"
                    )
                    self.load_recipes()
                else:
                    error_detail = r.json().get("detail", "Ошибка")
                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Количество должно быть числом!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def delete_recipe(self, recipe_id, product_name):
        """Удаление рецептуры"""
        reply = QMessageBox.question(
            self,
            "Удаление рецептуры",
            f"Вы уверены, что хотите удалить рецептуру для '{product_name}'?\nЭто действие необратимо!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                r = requests.delete(
                    f"{API_URL}/admin/recipes/{recipe_id}",
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 200:
                    QMessageBox.information(self, "Успех", f"Рецептура '{product_name}' удалена!")
                    self.load_recipes()
                else:
                    error_detail = r.json().get("detail", "Ошибка удаления")
                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def add_recipe_dialog(self):
        """Диалог добавления рецептуры"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить технологическую карту")
        dialog.setMinimumWidth(450)

        layout = QFormLayout(dialog)

        # Название изделия
        name_input = QLineEdit()
        name_input.setPlaceholderText("Например: Бокс №10")
        layout.addRow("Название изделия:", name_input)

        # Толщина материала
        thickness_combo = QComboBox()
        thickness_combo.addItems(["2мм", "1.5мм", "3мм"])
        thickness_combo.setEditable(True)
        layout.addRow("Толщина материала:", thickness_combo)

        # Количество изделий в одном листе
        blanks_input = QLineEdit()
        blanks_input.setPlaceholderText("Сколько изделий помещается в 1 листе")
        layout.addRow("Изделий в 1 листе:", blanks_input)

        # Пояснение
        info_label = QLabel("💡 Пример: в 1 листе 3мм помещается 6 изделий")
        info_label.setStyleSheet("color: #ffc107; font-style: italic;")
        layout.addRow(info_label)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            product_name = name_input.text().strip()
            thickness = thickness_combo.currentText()
            blanks_text = blanks_input.text().strip()

            if not all([product_name, thickness, blanks_text]):
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                blanks_per_sheet = int(blanks_text)
                if blanks_per_sheet <= 0:
                    raise ValueError

                payload = {
                    "product_name": product_name.lower(),
                    "material_thickness": thickness,
                    "blanks_per_sheet": blanks_per_sheet
                }

                r = requests.post(
                    f"{API_URL}/admin/recipes/add",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 201:
                    QMessageBox.information(
                        self, "Успех",
                        f"Рецептура добавлена!\n\n"
                        f"Изделие: {product_name}\n"
                        f"Толщина: {thickness}\n"
                        f"Изделий в листе: {blanks_per_sheet}\n\n"
                        f"Пример: при добавлении 10 изделий будет списано "
                        f"{math.ceil(10 / blanks_per_sheet)} листов"
                    )
                    self.load_recipes()
                else:
                    error_detail = r.json().get("detail", "Ошибка")
                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Количество должно быть числом!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def load_recipes(self):
        """Загрузка списка рецептур"""
        print(f"🔄 Загрузка рецептур...")

        try:
            r = requests.get(f"{API_URL}/recipes/list")
            print(f"📡 Статус ответа: {r.status_code}")

            if r.status_code == 200:
                recipes = r.json()
                print(f"✅ Получено рецептур: {len(recipes)}")

                self.recipes_table.setRowCount(len(recipes))

                for i, recipe in enumerate(recipes):
                    print(f"  [{i}] {recipe['product_name']}: {recipe['material_thickness']}, "
                          f"{recipe['blanks_per_sheet']} шт/лист")

                    self.recipes_table.setItem(i, 0, QTableWidgetItem(str(recipe['id'])))
                    self.recipes_table.setItem(i, 1, QTableWidgetItem(recipe['product_name']))
                    self.recipes_table.setItem(i, 2, QTableWidgetItem(recipe['material_thickness']))
                    self.recipes_table.setItem(i, 3, QTableWidgetItem(str(recipe['blanks_per_sheet'])))

                    # === СОЗДАЁМ КОНТЕЙНЕР С КНОПКАМИ ===
                    container = QWidget()
                    btn_layout = QHBoxLayout(container)
                    btn_layout.setContentsMargins(2, 2, 2, 2)
                    btn_layout.setSpacing(5)

                    # Кнопка Изменить
                    edit_btn = QPushButton("✏️")
                    edit_btn.setToolTip("Изменить рецептуру")
                    edit_btn.setMinimumWidth(40)
                    edit_btn.setStyleSheet("""
                        QPushButton {
                            background: #28a745;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                            padding: 5px;
                        }
                        QPushButton:hover {
                            background: #218838;
                        }
                    """)
                    edit_btn.clicked.connect(
                        lambda checked, rid=recipe['id'], rname=recipe['product_name'],
                               rthickness=recipe['material_thickness'],
                               rblanks=recipe['blanks_per_sheet']:
                        self.edit_recipe_dialog(rid, rname, rthickness, rblanks)
                    )

                    # Кнопка Удалить
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setToolTip("Удалить рецептуру")
                    delete_btn.setMinimumWidth(40)
                    delete_btn.setStyleSheet("""
                        QPushButton {
                            background: #dc3545;
                            color: white;
                            border: none;
                            border-radius: 5px;
                            font-size: 14px;
                            padding: 5px;
                        }
                        QPushButton:hover {
                            background: #c82333;
                        }
                    """)
                    delete_btn.clicked.connect(
                        lambda checked, rid=recipe['id'], rname=recipe['product_name']:
                        self.delete_recipe(rid, rname)
                    )

                    btn_layout.addWidget(edit_btn)
                    btn_layout.addWidget(delete_btn)

                    # Добавляем контейнер в ячейку
                    self.recipes_table.setCellWidget(i, 4, container)
                    print(f"  ✅ Кнопки добавлены в строку {i}")

                # Скрываем ID
                self.recipes_table.setColumnHidden(0, True)

                # Устанавливаем ширину колонки действий
                self.recipes_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
                self.recipes_table.setColumnWidth(4, 120)

                print(f"✅ Загружено рецептур: {len(recipes)}")
            else:
                print(f"❌ Ошибка загрузки: {r.status_code}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить: {r.status_code}")
        except Exception as e:
            print(f" Исключение: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", str(e))

    def add_recipe_dialog(self):
        """Диалог добавления рецептуры"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить технологическую карту")
        dialog.setMinimumWidth(450)

        layout = QFormLayout(dialog)

        # Название изделия
        name_input = QLineEdit()
        name_input.setPlaceholderText("Например: Бокс №10")
        layout.addRow("Название изделия:", name_input)

        # Толщина материала
        thickness_combo = QComboBox()
        thickness_combo.addItems(["2мм", "1.5мм", "3мм"])
        thickness_combo.setEditable(True)
        layout.addRow("Толщина материала:", thickness_combo)

        # Заготовок в одном листе
        blanks_input = QLineEdit()
        blanks_input.setPlaceholderText("Сколько заготовок помещается в 1 листе")
        layout.addRow("Заготовок в 1 листе:", blanks_input)

        # Пояснение
        info_label = QLabel("💡 Пример: в 1 листе 3мм помещается 6 заготовок")
        info_label.setStyleSheet("color: #ffc107; font-style: italic;")
        layout.addRow(info_label)

        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            product_name = name_input.text().strip()
            thickness = thickness_combo.currentText()
            blanks_text = blanks_input.text().strip()

            if not all([product_name, thickness, blanks_text]):
                QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
                return

            try:
                blanks_per_sheet = int(blanks_text)
                if blanks_per_sheet <= 0:
                    raise ValueError

                payload = {
                    "product_name": product_name.lower(),
                    "material_thickness": thickness,
                    "blanks_per_sheet": blanks_per_sheet
                }

                r = requests.post(
                    f"{API_URL}/admin/recipes/add",
                    json=payload,
                    headers={"Authorization": f"Bearer {self.token}"}
                )

                if r.status_code == 201:
                    QMessageBox.information(
                        self, "Успех",
                        f"Рецептура добавлена!\n\n"
                        f"Изделие: {product_name}\n"
                        f"Толщина: {thickness}\n"
                        f"Заготовок в листе: {blanks_per_sheet}"
                    )
                    self.load_recipes()
                else:
                    error_detail = r.json().get("detail", "Ошибка")
                    QMessageBox.critical(self, "Ошибка", str(error_detail))

            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Количество должно быть числом!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))


if __name__ == "__main__":
    try:
        mutex = ctypes.windll.kernel32.CreateMutexW(None, False, MUTEX_NAME)
        if ctypes.windll.kernel32.GetLastError() == 183:
            print("Приложение уже запущено!")
            sys.exit(1)

        print("Запуск приложения...")
        app = QApplication(sys.argv)
        print("QApplication создан")
        app.setStyle("Fusion")
        window = ShiftApp()
        print("Окно создано, показываем...")
        window.show()
        print("Запуск цикла событий...")
        sys.exit(app.exec())
    except Exception as e:
        print(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback

        traceback.print_exc()
        input("Нажмите Enter для выхода...")